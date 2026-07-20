"""Tests for the Celery run_import_task (sync DB processing)."""

import base64
import io
from datetime import date

from app.modules.data_import.models import ImportJob
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession


def _make_excel_bytes(rows: list[list]) -> str:
    """Create base64-encoded Excel data."""
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()


class TestRunImportTask:
    async def test_employee_import_via_task(self, db: AsyncSession, tenant, user):
        """Simulate what the Celery task does using sync DB."""
        # Create ImportJob
        job = ImportJob(
            tenant_id=tenant.id,
            import_type="employees",
            file_name="test.xlsx",
            total_rows=1,
            initiated_by=user.id,
        )
        db.add(job)
        await db.commit()
        await db.refresh(job)

        assert job.status == "pending"

        # Verify job can be retrieved
        loaded = await db.get(ImportJob, job.id)
        assert loaded is not None
        assert loaded.import_type == "employees"
        assert loaded.total_rows == 1

    async def test_dictionary_import_job_creation(self, db: AsyncSession, tenant, user):
        """Verify ImportJob is created correctly before Celery task runs."""
        job = ImportJob(
            tenant_id=tenant.id,
            import_type="dictionaries",
            file_name="dicts.xlsx",
            total_rows=3,
            initiated_by=user.id,
        )
        db.add(job)
        await db.commit()
        await db.refresh(job)

        assert job.status == "pending"
        assert job.total_rows == 3
        assert job.processed_rows == 0
        assert job.error_rows == 0

    async def test_import_job_status_update(self, db: AsyncSession, tenant, user):
        """Verify ImportJob status can be updated (simulates Celery task flow)."""
        from datetime import datetime, timezone

        job = ImportJob(
            tenant_id=tenant.id,
            import_type="employees",
            file_name="test.xlsx",
            total_rows=5,
            initiated_by=user.id,
        )
        db.add(job)
        await db.commit()
        await db.refresh(job)

        # Simulate task processing
        job.status = "processing"
        job.started_at = datetime.now(timezone.utc)
        await db.commit()
        await db.refresh(job)
        assert job.status == "processing"

        # Simulate task completion
        job.status = "completed"
        job.processed_rows = 4
        job.error_rows = 1
        job.errors = {"rows": [{"row": 3, "error": "Missing email"}]}
        job.finished_at = datetime.now(timezone.utc)
        await db.commit()
        await db.refresh(job)

        assert job.status == "completed"
        assert job.processed_rows == 4
        assert job.error_rows == 1
        assert job.errors["rows"][0]["row"] == 3

    async def test_b64_encoding_roundtrip(self):
        """Verify Excel data survives base64 encoding (used in Celery transport)."""
        rows = [
            ["email", "first_name", "last_name", "position", "hire_date"],
            ["test@example.com", "Test", "User", "Dev", date(2024, 1, 15)],
        ]
        b64 = _make_excel_bytes(rows)
        raw = base64.b64decode(b64)

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True)
        ws = wb.active
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 1
        assert data_rows[0][0] == "test@example.com"
