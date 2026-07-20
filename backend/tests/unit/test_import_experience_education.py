"""Tests for GF1/GF2: Import support for work experience, education, and courses.

With async import (I3), start_import() queues a Celery task and returns immediately.
These tests verify that the job is created correctly and the Celery task is queued.
The actual row processing logic is tested via test_celery_import_task.py.
"""

import base64
from datetime import date
from io import BytesIO
from unittest.mock import MagicMock, patch

from app.modules.data_import import service as import_service
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession


def _make_employee_excel(*rows, headers=None) -> bytes:
    wb = Workbook()
    ws = wb.active
    if headers is None:
        headers = [
            "email",
            "first_name",
            "last_name",
            "position",
            "hire_date",
            "work_exp_title",
            "work_exp_role",
            "work_exp_start",
            "work_exp_end",
            "edu_institution",
            "edu_degree",
            "edu_field",
            "edu_start",
            "edu_end",
            "course_title",
            "course_provider",
            "course_completed_date",
        ]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class _FakeUpload:
    def __init__(self, data: bytes, filename: str = "test.xlsx"):
        self._data = data
        self.filename = filename

    async def read(self) -> bytes:
        return self._data


class TestImportWorkExperience:
    @patch("app.modules.data_import.tasks.run_import_task")
    async def test_import_with_work_experience(
        self, mock_task, db: AsyncSession, tenant, user
    ):
        mock_task.apply_async = MagicMock()
        data = _make_employee_excel(
            [
                "wx@test.com",
                "Work",
                "Exp",
                "Dev",
                date(2023, 1, 1),
                "Project Alpha",
                "Lead Dev",
                date(2022, 1, 1),
                date(2022, 12, 31),
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ],
        )
        result = await import_service.start_import(
            db, tenant.id, user.id, "employees", _FakeUpload(data)
        )
        assert result["status"] == "pending"
        assert result["total_rows"] == 1
        mock_task.apply_async.assert_called_once()

        # Verify base64-encoded file data was passed to Celery
        call_args = mock_task.apply_async.call_args.kwargs["args"]
        b64_data = call_args[4]
        raw = base64.b64decode(b64_data)
        wb = load_workbook(BytesIO(raw), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1
        assert rows[0][0] == "wx@test.com"

    @patch("app.modules.data_import.tasks.run_import_task")
    async def test_import_without_optional_columns(
        self, mock_task, db: AsyncSession, tenant, user
    ):
        """Basic employee import without any optional columns still works."""
        mock_task.apply_async = MagicMock()
        data = _make_employee_excel(
            ["basic@test.com", "Basic", "User", "Tester", date(2024, 1, 1)],
        )
        result = await import_service.start_import(
            db, tenant.id, user.id, "employees", _FakeUpload(data)
        )
        assert result["status"] == "pending"
        assert result["total_rows"] == 1
        mock_task.apply_async.assert_called_once()

    @patch("app.modules.data_import.tasks.run_import_task")
    async def test_work_exp_skipped_without_start_date(
        self, mock_task, db: AsyncSession, tenant, user
    ):
        """File data is passed to Celery task correctly even with sparse optional columns."""
        mock_task.apply_async = MagicMock()
        data = _make_employee_excel(
            [
                "nodate@test.com",
                "No",
                "Date",
                "Dev",
                date(2023, 1, 1),
                "Project Beta",
                "Dev",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ],
        )
        result = await import_service.start_import(
            db, tenant.id, user.id, "employees", _FakeUpload(data)
        )
        assert result["status"] == "pending"
        assert result["total_rows"] == 1
        mock_task.apply_async.assert_called_once()


class TestImportEducation:
    @patch("app.modules.data_import.tasks.run_import_task")
    async def test_import_with_education(
        self, mock_task, db: AsyncSession, tenant, user
    ):
        mock_task.apply_async = MagicMock()
        data = _make_employee_excel(
            [
                "edu@test.com",
                "Edu",
                "User",
                "Dev",
                date(2023, 1, 1),
                None,
                None,
                None,
                None,
                "MIT",
                "Master",
                "CS",
                date(2018, 9, 1),
                date(2020, 6, 1),
                None,
                None,
                None,
            ],
        )
        result = await import_service.start_import(
            db, tenant.id, user.id, "employees", _FakeUpload(data)
        )
        assert result["status"] == "pending"
        assert result["total_rows"] == 1

    @patch("app.modules.data_import.tasks.run_import_task")
    async def test_import_with_course(self, mock_task, db: AsyncSession, tenant, user):
        mock_task.apply_async = MagicMock()
        data = _make_employee_excel(
            [
                "course@test.com",
                "Course",
                "User",
                "Dev",
                date(2023, 1, 1),
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "AWS Cert",
                "Amazon",
                date(2024, 3, 15),
            ],
        )
        result = await import_service.start_import(
            db, tenant.id, user.id, "employees", _FakeUpload(data)
        )
        assert result["status"] == "pending"
        assert result["total_rows"] == 1

    @patch("app.modules.data_import.tasks.run_import_task")
    async def test_import_all_optional_columns(
        self, mock_task, db: AsyncSession, tenant, user
    ):
        """Import with all optional columns filled — job created and task queued."""
        mock_task.apply_async = MagicMock()
        data = _make_employee_excel(
            [
                "full@test.com",
                "Full",
                "Import",
                "Dev",
                date(2023, 1, 1),
                "Project X",
                "Lead",
                date(2022, 1, 1),
                date(2022, 12, 31),
                "Stanford",
                "PhD",
                "AI",
                date(2015, 9, 1),
                date(2020, 6, 1),
                "ML Course",
                "Coursera",
                date(2023, 6, 1),
            ],
        )
        result = await import_service.start_import(
            db, tenant.id, user.id, "employees", _FakeUpload(data)
        )
        assert result["status"] == "pending"
        assert result["total_rows"] == 1
        mock_task.apply_async.assert_called_once()

        # Verify the import type was passed correctly
        call_args = mock_task.apply_async.call_args.kwargs["args"]
        assert call_args[3] == "employees"
