"""Dedicated suite for ``app.modules.analytics.service`` (review item [25]).

Pins the observable output shape of every public report function. All data
is scoped to the test's own tenant so the suite stays deterministic on the
shared test DB.
"""

import io
import uuid
from datetime import date, datetime, timezone

import pytest_asyncio
from app.core.security import hash_password
from app.modules.analytics import service as analytics_service
from app.modules.assessment.models import (
    CPA,
    PDP,
    Assessment,
    AssessmentResult,
    AssessmentStatus,
    AssessmentType,
    PDPVersion,
)
from app.modules.auth.models import User
from app.modules.company.models import Division, SpecializationDivision
from app.modules.competence.models import Competence, CompetenceGroup
from app.modules.dictionary.models import DictionaryItem
from app.modules.employee.models import Compensation, Employee
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession


@pytest_asyncio.fixture
async def status_done(db: AsyncSession):
    result = await db.execute(
        select(AssessmentStatus).where(AssessmentStatus.code == "done")
    )
    s = result.scalars().first()
    if not s:
        s = AssessmentStatus(code="done", title="Done", sequence=6)
        db.add(s)
        await db.commit()
        await db.refresh(s)
    return s


@pytest_asyncio.fixture
async def type_self(db: AsyncSession):
    result = await db.execute(
        select(AssessmentType).where(AssessmentType.code == "self")
    )
    t = result.scalars().first()
    if not t:
        t = AssessmentType(code="self", title="Self assessment")
        db.add(t)
        await db.commit()
        await db.refresh(t)
    return t


async def _make_employee(db: AsyncSession, tenant, *, name: str) -> Employee:
    suffix = uuid.uuid4().hex[:8]
    u = User(
        email=f"analytics-{suffix}@test.com",
        password_hash=hash_password("testpass123"),
        first_name=name,
        last_name="Analyticstest",
        tenant_id=tenant.id,
        email_verified_at=datetime.now(timezone.utc),
    )
    db.add(u)
    await db.flush()
    emp = Employee(
        user_id=u.id,
        tenant_id=tenant.id,
        position_title="Engineer",
        hire_date=date(2024, 1, 15),
    )
    db.add(emp)
    await db.flush()
    return emp


async def _make_assessment(
    db: AsyncSession,
    tenant,
    emp: Employee,
    status: AssessmentStatus,
    a_type: AssessmentType,
    *,
    initiator_id: uuid.UUID,
    **extra,
) -> Assessment:
    a = Assessment(
        tenant_id=tenant.id,
        title=f"A-{uuid.uuid4().hex[:6]}",
        employee_id=emp.id,
        type_id=a_type.id,
        status_id=status.id,
        initiator_id=initiator_id,
        **extra,
    )
    db.add(a)
    await db.flush()
    return a


async def _make_competence(db: AsyncSession, tenant) -> Competence:
    group = CompetenceGroup(title=f"G-{uuid.uuid4().hex[:6]}", tenant_id=tenant.id)
    db.add(group)
    await db.flush()
    comp = Competence(
        title=f"C-{uuid.uuid4().hex[:6]}", group_id=group.id, tenant_id=tenant.id
    )
    db.add(comp)
    await db.flush()
    return comp


class TestAssessmentAndPdpStats:
    async def test_assessment_stats_counts_and_avg(
        self, db: AsyncSession, tenant, user, status_done, type_self
    ):
        emp = await _make_employee(db, tenant, name="Stats")
        a1 = await _make_assessment(
            db, tenant, emp, status_done, type_self, initiator_id=user.id
        )
        await _make_assessment(
            db, tenant, emp, status_done, type_self, initiator_id=user.id
        )
        comp = await _make_competence(db, tenant)
        db.add(
            AssessmentResult(assessment_id=a1.id, competence_id=comp.id, avg_score=4.0)
        )
        db.add(
            AssessmentResult(assessment_id=a1.id, competence_id=comp.id, avg_score=3.0)
        )
        await db.commit()

        stats = await analytics_service.assessment_stats(db, tenant.id)
        assert stats["total"] == 2
        assert stats["by_status"] == {"done": 2}
        assert stats["avg_score"] == 3.5

    async def test_assessment_stats_empty_tenant(self, db: AsyncSession, tenant):
        stats = await analytics_service.assessment_stats(db, tenant.id)
        assert stats == {"total": 0, "by_status": {}, "avg_score": None}

    async def test_pdp_stats(self, db: AsyncSession, tenant, user):
        emp = await _make_employee(db, tenant, name="Pdp")
        db.add(
            PDP(
                tenant_id=tenant.id,
                title="PDP-1",
                employee_id=emp.id,
                author_id=user.id,
                status="in_progress",
                total_progress=40,
            )
        )
        db.add(
            PDP(
                tenant_id=tenant.id,
                title="PDP-2",
                employee_id=emp.id,
                author_id=user.id,
                status="done",
                total_progress=100,
            )
        )
        await db.commit()

        stats = await analytics_service.pdp_stats(db, tenant.id)
        assert stats["total"] == 2
        assert stats["by_status"] == {"in_progress": 1, "done": 1}
        assert stats["avg_progress"] == 70.0


class TestExportAssessmentsXlsx:
    async def test_export_contains_rows(
        self, db: AsyncSession, tenant, user, status_done, type_self
    ):
        emp = await _make_employee(db, tenant, name="Export")
        a = await _make_assessment(
            db, tenant, emp, status_done, type_self, initiator_id=user.id
        )
        await db.commit()

        response = await analytics_service.export_assessments_xlsx(db, tenant.id)
        assert response.headers["content-disposition"].endswith("assessments.xlsx")

        body = b"".join([chunk async for chunk in response.body_iterator])
        wb = load_workbook(io.BytesIO(body))
        ws = wb["Assessments"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("ID", "Title", "Employee ID", "Status", "Created At")
        assert len(rows) == 2
        assert rows[1][0] == str(a.id)
        assert rows[1][1] == a.title
        assert rows[1][3] == "done"


class TestCompensationBenchmark:
    async def test_benchmark_by_grade_and_specialization(
        self, db: AsyncSession, tenant, user, status_done, type_self
    ):
        emp = await _make_employee(db, tenant, name="Bench")
        grade = DictionaryItem(type="grade", title="Senior", tenant_id=tenant.id)
        spec = DictionaryItem(
            type="specialization", title="Backend", tenant_id=tenant.id
        )
        db.add_all([grade, spec])
        await db.flush()
        await _make_assessment(
            db,
            tenant,
            emp,
            status_done,
            type_self,
            initiator_id=user.id,
            grade_id=grade.id,
            specialization_id=spec.id,
        )
        db.add(
            Compensation(
                employee_id=emp.id,
                tenant_id=tenant.id,
                type="salary",
                amount=300000,
                currency="USD",
                effective_date=date(2024, 1, 1),
            )
        )
        await db.commit()

        result = await analytics_service.compensation_benchmark(db, tenant.id)
        assert result["by_grade"] == [
            {
                "grade_id": str(grade.id),
                "grade_title": "Senior",
                "avg_salary": 300000.0,
                "count": 1,
            }
        ]
        assert result["by_specialization"] == [
            {
                "specialization_id": str(spec.id),
                "specialization_title": "Backend",
                "avg_salary": 300000.0,
                "count": 1,
            }
        ]

    async def test_benchmark_empty(self, db: AsyncSession, tenant):
        result = await analytics_service.compensation_benchmark(db, tenant.id)
        assert result == {"by_grade": [], "by_specialization": []}


class TestDivisionSpecializationMatrix:
    async def test_matrix_cells_and_headcount(self, db: AsyncSession, tenant):
        emp = await _make_employee(db, tenant, name="Matrix")
        div = Division(tenant_id=tenant.id, name="Engineering")
        spec = DictionaryItem(type="specialization", title="QA", tenant_id=tenant.id)
        db.add_all([div, spec])
        await db.flush()
        emp.division_id = div.id
        db.add(
            SpecializationDivision(
                tenant_id=tenant.id, division_id=div.id, specialization_id=spec.id
            )
        )
        await db.commit()

        result = await analytics_service.division_specialization_matrix(db, tenant.id)
        assert result["divisions"] == [{"id": str(div.id), "name": "Engineering"}]
        assert result["specializations"] == [{"id": str(spec.id), "title": "QA"}]
        assert result["cells"] == [
            {
                "division_id": str(div.id),
                "specialization_id": str(spec.id),
                "headcount": 1,
                "active": True,
            }
        ]

    async def test_unmapped_division_produces_no_cell(self, db: AsyncSession, tenant):
        db.add(Division(tenant_id=tenant.id, name="Sales"))
        await db.commit()

        result = await analytics_service.division_specialization_matrix(db, tenant.id)
        assert len(result["divisions"]) == 1
        assert result["cells"] == []


class TestCompareCpaRounds:
    async def test_deltas_per_employee(
        self, db: AsyncSession, tenant, user, status_done, type_self
    ):
        emp = await _make_employee(db, tenant, name="Rounds")
        comp = await _make_competence(db, tenant)
        cpa1 = CPA(
            tenant_id=tenant.id, title="Q1", author_id=user.id, type_id=type_self.id
        )
        cpa2 = CPA(
            tenant_id=tenant.id, title="Q2", author_id=user.id, type_id=type_self.id
        )
        db.add_all([cpa1, cpa2])
        await db.flush()
        a1 = await _make_assessment(
            db,
            tenant,
            emp,
            status_done,
            type_self,
            initiator_id=user.id,
            cpa_id=cpa1.id,
        )
        a2 = await _make_assessment(
            db,
            tenant,
            emp,
            status_done,
            type_self,
            initiator_id=user.id,
            cpa_id=cpa2.id,
        )
        db.add(
            AssessmentResult(assessment_id=a1.id, competence_id=comp.id, avg_score=3.0)
        )
        db.add(
            AssessmentResult(assessment_id=a2.id, competence_id=comp.id, avg_score=4.5)
        )
        await db.commit()

        result = await analytics_service.compare_cpa_rounds(
            db, tenant.id, cpa1.id, cpa2.id
        )
        assert result["round_1"] == {"id": str(cpa1.id), "title": "Q1"}
        assert result["round_2"] == {"id": str(cpa2.id), "title": "Q2"}
        assert len(result["comparisons"]) == 1
        row = result["comparisons"][0]
        assert row["employee_id"] == str(emp.id)
        assert row["employee_name"] == "Rounds Analyticstest"
        assert row["round_1_score"] == 3.0
        assert row["round_2_score"] == 4.5
        assert row["delta"] == 1.5

    async def test_employee_missing_in_one_round_has_null_delta(
        self, db: AsyncSession, tenant, user, status_done, type_self
    ):
        emp = await _make_employee(db, tenant, name="OneRound")
        comp = await _make_competence(db, tenant)
        cpa1 = CPA(
            tenant_id=tenant.id, title="R1", author_id=user.id, type_id=type_self.id
        )
        cpa2 = CPA(
            tenant_id=tenant.id, title="R2", author_id=user.id, type_id=type_self.id
        )
        db.add_all([cpa1, cpa2])
        await db.flush()
        a1 = await _make_assessment(
            db,
            tenant,
            emp,
            status_done,
            type_self,
            initiator_id=user.id,
            cpa_id=cpa1.id,
        )
        db.add(
            AssessmentResult(assessment_id=a1.id, competence_id=comp.id, avg_score=2.0)
        )
        await db.commit()

        result = await analytics_service.compare_cpa_rounds(
            db, tenant.id, cpa1.id, cpa2.id
        )
        row = result["comparisons"][0]
        assert row["round_1_score"] == 2.0
        assert row["round_2_score"] is None
        assert row["delta"] is None


class TestPdpProgressTimeline:
    async def test_timeline_from_version_snapshots(
        self, db: AsyncSession, tenant, user
    ):
        emp = await _make_employee(db, tenant, name="Timeline")
        pdp = PDP(
            tenant_id=tenant.id,
            title="Growth",
            employee_id=emp.id,
            author_id=user.id,
            status="in_progress",
        )
        db.add(pdp)
        await db.flush()
        db.add(
            PDPVersion(
                pdp_id=pdp.id,
                version_number=1,
                status="draft",
                snapshot={
                    "items": [{"is_passed": False}, {"is_passed": False}],
                },
            )
        )
        db.add(
            PDPVersion(
                pdp_id=pdp.id,
                version_number=2,
                status="in_progress",
                snapshot={
                    "items": [{"is_passed": True}, {"is_passed": False}],
                },
            )
        )
        await db.commit()

        timeline = await analytics_service.pdp_progress_timeline(db, tenant.id, pdp.id)
        assert [t["version"] for t in timeline] == [1, 2]
        assert timeline[0]["progress"] == 0
        assert timeline[1]["progress"] == 50
        assert timeline[1]["total_items"] == 2
        assert timeline[1]["passed_items"] == 1

    async def test_timeline_is_tenant_scoped(self, db: AsyncSession, tenant, user):
        """A pdp_id from another tenant must yield an empty timeline."""
        emp = await _make_employee(db, tenant, name="Scoped")
        pdp = PDP(
            tenant_id=tenant.id,
            title="Private",
            employee_id=emp.id,
            author_id=user.id,
            status="draft",
        )
        db.add(pdp)
        await db.flush()
        db.add(
            PDPVersion(
                pdp_id=pdp.id,
                version_number=1,
                status="draft",
                snapshot={"items": []},
            )
        )
        await db.commit()

        foreign_tenant_id = uuid.uuid4()
        timeline = await analytics_service.pdp_progress_timeline(
            db, foreign_tenant_id, pdp.id
        )
        assert timeline == []
