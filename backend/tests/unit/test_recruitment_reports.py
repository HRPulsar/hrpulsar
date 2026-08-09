"""Unit tests for R4a — XLSX consolidated reports."""

from __future__ import annotations

import base64
import io
import uuid
from unittest.mock import patch

import pytest
from app.modules.recruitment import service
from app.modules.recruitment.models import ConsolidatedReport
from app.modules.recruitment.report_xlsx import render_report_xlsx
from app.modules.recruitment.schemas import (
    CandidateCreate,
    CandidateVacancyCreate,
    ReportGenerateRequest,
    VacancyCreate,
)
from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession


async def _make_vacancy(db: AsyncSession, tenant, user, title: str = "V"):
    vac = await service.create_vacancy(
        db,
        tenant.id,
        user.id,
        VacancyCreate(title=f"{title}-{uuid.uuid4().hex[:5]}"),
    )
    return vac


async def _make_candidate(db: AsyncSession, tenant, user):
    return await service.create_candidate(
        db,
        tenant.id,
        user.id,
        CandidateCreate(
            first_name=f"Cand{uuid.uuid4().hex[:4]}",
            last_name=f"Last{uuid.uuid4().hex[:4]}",
            email=f"{uuid.uuid4().hex[:6]}@example.com",
        ),
    )


# ---------------------------------------------------------------------------
# Enqueue / list / get / delete
# ---------------------------------------------------------------------------


class TestReportEnqueue:
    async def test_enqueue_creates_pending_export_and_dispatches_task(
        self, db: AsyncSession, tenant, user
    ) -> None:
        vac = await _make_vacancy(db, tenant, user)

        with patch(
            "app.modules.recruitment.tasks.generate_report_task.delay"
        ) as mock_delay:
            mock_delay.return_value.id = "fake-task-id"
            result = await service.enqueue_report(
                db,
                tenant.id,
                user.id,
                uuid.UUID(str(vac["id"])),
                ReportGenerateRequest(sections=["summary_ranking"]),
            )

        assert result["task_id"] == "fake-task-id"
        export = (
            await db.execute(
                select(ConsolidatedReport).where(
                    ConsolidatedReport.id == result["export_id"]
                )
            )
        ).scalar_one()
        assert export.status == "pending"
        assert export.sections == ["summary_ranking"]
        assert export.generated_by == user.id
        mock_delay.assert_called_once_with(str(result["export_id"]), str(tenant.id))

    async def test_enqueue_uses_default_sections_when_none_given(
        self, db: AsyncSession, tenant, user
    ) -> None:
        vac = await _make_vacancy(db, tenant, user)
        with patch(
            "app.modules.recruitment.tasks.generate_report_task.delay"
        ) as mock_delay:
            mock_delay.return_value.id = "tid"
            result = await service.enqueue_report(
                db,
                tenant.id,
                user.id,
                uuid.UUID(str(vac["id"])),
                ReportGenerateRequest(),
            )
        export = (
            await db.execute(
                select(ConsolidatedReport).where(
                    ConsolidatedReport.id == result["export_id"]
                )
            )
        ).scalar_one()
        assert "summary_ranking" in (export.sections or [])
        assert "competency_matrix" in (export.sections or [])

    async def test_empty_candidate_selection_is_rejected(
        self, db: AsyncSession, tenant, user
    ) -> None:
        """An empty whitelist used to be indistinguishable from "no
        whitelist" — the truthiness check dropped it and the worker
        reported on every candidate on the vacancy, the opposite of the
        Finalists/Custom scope the user picked."""
        vac = await _make_vacancy(db, tenant, user)
        with pytest.raises(HTTPException) as exc:
            await service.enqueue_report(
                db,
                tenant.id,
                user.id,
                uuid.UUID(str(vac["id"])),
                ReportGenerateRequest(candidate_vacancy_ids=[]),
            )
        assert exc.value.status_code == 400
        assert exc.value.code == "report_candidate_selection_empty"

    async def test_omitted_candidate_selection_still_means_all_active(
        self, db: AsyncSession, tenant, user
    ) -> None:
        vac = await _make_vacancy(db, tenant, user)
        with patch(
            "app.modules.recruitment.tasks.generate_report_task.delay"
        ) as mock_delay:
            mock_delay.return_value.id = "tid"
            result = await service.enqueue_report(
                db,
                tenant.id,
                user.id,
                uuid.UUID(str(vac["id"])),
                ReportGenerateRequest(),
            )
        export = (
            await db.execute(
                select(ConsolidatedReport).where(
                    ConsolidatedReport.id == result["export_id"]
                )
            )
        ).scalar_one()
        assert "candidate_vacancy_ids" not in (export.report_data or {})

    async def test_enqueue_unknown_vacancy_returns_404(
        self, db: AsyncSession, tenant, user
    ) -> None:
        with pytest.raises(HTTPException) as exc:
            await service.enqueue_report(
                db,
                tenant.id,
                user.id,
                uuid.uuid4(),
                ReportGenerateRequest(),
            )
        assert exc.value.status_code == 404


class TestReportListAndGet:
    async def test_list_and_get(self, db: AsyncSession, tenant, user) -> None:
        vac = await _make_vacancy(db, tenant, user)
        with patch(
            "app.modules.recruitment.tasks.generate_report_task.delay"
        ) as mock_delay:
            mock_delay.return_value.id = "tid"
            res = await service.enqueue_report(
                db,
                tenant.id,
                user.id,
                uuid.UUID(str(vac["id"])),
                ReportGenerateRequest(sections=["summary_ranking"]),
            )

        items, total = await service.list_reports(
            db, tenant.id, vacancy_id=uuid.UUID(str(vac["id"]))
        )
        assert total == 1
        assert items[0]["status"] == "pending"
        assert items[0]["sections"] == ["summary_ranking"]
        assert items[0]["requested_by_name"]

        single = await service.get_report(
            db, tenant.id, uuid.UUID(str(res["export_id"]))
        )
        assert single["id"] == res["export_id"]
        assert single["download_url"] is None  # still pending

    async def test_get_unknown_returns_404(self, db: AsyncSession, tenant) -> None:
        with pytest.raises(HTTPException) as exc:
            await service.get_report(db, tenant.id, uuid.uuid4())
        assert exc.value.status_code == 404

    async def test_cross_tenant_access_returns_404(
        self, db: AsyncSession, tenant, user
    ) -> None:
        vac = await _make_vacancy(db, tenant, user)
        with patch(
            "app.modules.recruitment.tasks.generate_report_task.delay"
        ) as mock_delay:
            mock_delay.return_value.id = "tid"
            res = await service.enqueue_report(
                db,
                tenant.id,
                user.id,
                uuid.UUID(str(vac["id"])),
                ReportGenerateRequest(),
            )
        other_tenant_id = uuid.uuid4()
        with pytest.raises(HTTPException) as exc:
            await service.get_report(
                db, other_tenant_id, uuid.UUID(str(res["export_id"]))
            )
        assert exc.value.status_code == 404

    async def test_delete_report_clears_row(
        self, db: AsyncSession, tenant, user
    ) -> None:
        vac = await _make_vacancy(db, tenant, user)
        with patch(
            "app.modules.recruitment.tasks.generate_report_task.delay"
        ) as mock_delay:
            mock_delay.return_value.id = "tid"
            res = await service.enqueue_report(
                db,
                tenant.id,
                user.id,
                uuid.UUID(str(vac["id"])),
                ReportGenerateRequest(),
            )
        await service.delete_report(db, tenant.id, uuid.UUID(str(res["export_id"])))
        items, total = await service.list_reports(
            db, tenant.id, vacancy_id=uuid.UUID(str(vac["id"]))
        )
        assert total == 0


# ---------------------------------------------------------------------------
# XLSX renderer
# ---------------------------------------------------------------------------


class TestReportXlsxRenderer:
    def _vacancy_payload(self) -> dict:
        return {
            "title": "Senior Backend Engineer",
            "status": "open",
            "specialization_title": "Engineering",
            "grade_title": "Senior",
            "division_name": "Platform",
            "location": "Remote",
            "employment_type": "full_time",
            "salary_min": 100000,
            "salary_max": 140000,
            "salary_currency": "USD",
            "owner_name": "Alice Recruiter",
            "created_at": "2026-05-09T10:00:00",
            "description": "Hires only",
        }

    def test_render_emits_four_sheet_layout(self) -> None:
        """HRP-268 — always 4 sheets: Summary, Matrix, one Detail per
        candidate, and the Incomplete-data tail. Recruiter audience
        keeps the raw process-findings text on the Detail sheet."""
        cand_ids = [uuid.uuid4(), uuid.uuid4()]
        summary_rows = [
            {
                "candidate_name": "John Doe",
                "profile_summary": "10y backend, async stack",
                "manager_score_text": "8/10 (80%)",
                "ai_score_text": "9/10 (90%)",
                "data_readiness": "Complete",
                "recommendation": "Recommended",
            },
            {
                "candidate_name": "Mary Smith",
                "profile_summary": "5y Python + DS",
                "manager_score_text": "6/10 (60%)",
                "ai_score_text": "—",
                "data_readiness": "AI missing",
                "recommendation": "Additional check",
            },
        ]
        matrix = {
            "candidates": [
                {"id": cand_ids[0], "name": "John Doe"},
                {"id": cand_ids[1], "name": "Mary Smith"},
            ],
            "competences": [
                {
                    "id": uuid.uuid4(),
                    "name": "Python",
                    "group": "Hard",
                    "criticality": "critical",
                    "cells": [
                        {
                            "candidate_id": cand_ids[0],
                            "manager_score": 4.0,
                            "ai_score": 4.5,
                            "ai_status": "ready",
                            "divergence": False,
                        },
                        {
                            "candidate_id": cand_ids[1],
                            "manager_score": 3.0,
                            "ai_score": None,
                            "ai_status": "missing",
                            "divergence": False,
                        },
                    ],
                },
            ],
            "max_score": 5.0,
        }
        details = [
            {
                "candidate_name": "John Doe",
                "position": "Senior",
                "status": "interview",
                "resume_summary": "Worked on async services",
                "scores": [
                    {
                        "competence_name": "Python",
                        "manager_score": 4.0,
                        "ai_score": 4.5,
                        "citations": ["I built async services"],
                        "reasoning": "Strong async background",
                    }
                ],
                "blind_spots": [],
                "process_findings": [
                    {
                        "finding_type": "leading_question",
                        "severity": "low",
                        "description": "Leading question on architecture.",
                        "positive_reframe": "Try open-ended architecture probe.",
                    }
                ],
                "red_flags": [],
                "verdict": {"verdict_summary": "Strong"},
            }
        ]
        incomplete = [
            {
                "candidate_name": "Mary Smith",
                "missing": "AI analysis missing",
                "action": "Launch the AI analysis (resume-only or full).",
            }
        ]

        xlsx_bytes = render_report_xlsx(
            vacancy=self._vacancy_payload(),
            summary_rows=summary_rows,
            matrix=matrix,
            details=details,
            incomplete=incomplete,
            branding={"tenant_name": "Acme"},
        )
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert wb.sheetnames == [
            "Summary",
            "Matrix",
            "Detail · John Doe",
            "Incomplete data",
        ]

        ws_summary = wb["Summary"]
        # HRP-522/523 — row 1 is the logo band, rows 2-4 are title,
        # position/date line and the completeness disclaimer, so the
        # header row lands on row 6.
        assert ws_summary.cell(row=6, column=1).value == "Candidate"
        assert ws_summary.cell(row=7, column=1).value == "John Doe"

        ws_matrix = wb["Matrix"]
        # Total row carries a SUM formula on each Manager / AI column.
        total_row = ws_matrix.max_row
        manager_total = ws_matrix.cell(row=total_row, column=4).value
        assert isinstance(manager_total, str) and manager_total.startswith("=SUM(")

        ws_detail = wb["Detail · John Doe"]
        flat = [
            ws_detail.cell(row=r, column=1).value
            for r in range(1, ws_detail.max_row + 1)
        ]
        # Recruiter audience keeps the raw process-findings section.
        assert "─── PROCESS FINDINGS ───" in flat

        ws_incomplete = wb["Incomplete data"]
        assert ws_incomplete.cell(row=5, column=1).value == "Mary Smith"

    def test_render_hides_process_findings_for_hiring_manager(self) -> None:
        """Hiring-manager audience swaps the raw process-findings
        section for a positive reframe of the same payload."""
        details = [
            {
                "candidate_name": "John Doe",
                "scores": [],
                "process_findings": [
                    {
                        "finding_type": "leading_question",
                        "severity": "low",
                        "description": "Leading question on architecture.",
                        "positive_reframe": "Try open-ended architecture probe.",
                    }
                ],
            }
        ]
        xlsx_bytes = render_report_xlsx(
            vacancy=self._vacancy_payload(),
            summary_rows=[],
            matrix={"candidates": [], "competences": []},
            details=details,
            audience="hiring_manager",
        )
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws_detail = wb["Detail · John Doe"]
        flat = [
            ws_detail.cell(row=r, column=1).value
            for r in range(1, ws_detail.max_row + 1)
        ]
        # The HM-side header replaces "Process findings" with a neutral
        # "Recommendations for the next interview" title.
        assert "─── PROCESS RECOMMENDATIONS FOR NEXT INTERVIEW ───" in flat
        assert "─── PROCESS FINDINGS ───" not in flat
        # Raw process-findings description must NOT leak into the
        # workbook for the HM audience — only the reframe.
        sheet_dump = " ".join(
            str(ws_detail.cell(row=r, column=c).value or "")
            for r in range(1, ws_detail.max_row + 1)
            for c in range(1, ws_detail.max_column + 1)
        )
        assert "Leading question on architecture" not in sheet_dump
        assert "Try open-ended architecture probe" in sheet_dump

    def test_render_emits_incomplete_sheet_even_when_empty(self) -> None:
        xlsx_bytes = render_report_xlsx(
            vacancy={"title": "x"},
            summary_rows=[],
            matrix={"candidates": [], "competences": []},
            details=[],
            incomplete=[],
        )
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert wb.sheetnames == ["Summary", "Matrix", "Incomplete data"]
        ws = wb["Incomplete data"]
        assert "All candidates have complete data" in str(
            ws.cell(row=5, column=1).value
        )

    def test_detail_sheet_structure(self) -> None:
        """HRP-525 — sheet named after the candidate, spec header, resume
        merged across A:E, Scores carrying the group, and every section
        present as a banner."""
        detail = {
            "candidate_name": "Anna Smirnova",
            "position": "Head of Maintenance",
            "status": "Interview 2",
            "resume_summary": "14 years in maintenance planning.",
            "scores": [
                {
                    "competence_group": "PROF: Processes",
                    "competence_name": "Tariff analysis",
                    "manager_score": 2.0,
                    "ai_score": 1.0,
                    "citations": ["I ran the tariff model"],
                    "reasoning": "Solid but narrow.",
                }
            ],
            "blind_spots": [
                {
                    "competence": "Budgeting",
                    "suggested_question": "Walk me through your last budget.",
                }
            ],
            "process_findings": [
                {
                    "finding_type": "leading_question",
                    "severity": "low",
                    "description": "Leading question on architecture.",
                    "positive_reframe": "Try an open-ended probe.",
                }
            ],
            "red_flags": [
                {
                    "flag_type": "gap",
                    "severity": "medium",
                    "description": "Unexplained 8-month gap.",
                }
            ],
            "verdict": {
                "verdict_summary": "Strong operator",
                "key_strength": "Process depth",
                "key_risk": "Narrow tooling",
                "recommendation_for_next_step": "Technical deep dive",
            },
        }
        wb = load_workbook(
            io.BytesIO(
                render_report_xlsx(
                    vacancy=self._vacancy_payload(),
                    summary_rows=[],
                    matrix={"candidates": [], "competences": []},
                    details=[detail],
                    incomplete=[],
                    sections=["detailed_analysis"],
                )
            )
        )
        assert wb.sheetnames == ["Detail · Anna Smirnova"]
        ws = wb["Detail · Anna Smirnova"]
        assert ws.cell(row=2, column=1).value == "Detailed analysis — Anna Smirnova"
        subtitle = str(ws.cell(row=3, column=1).value)
        assert "Head of Maintenance" in subtitle
        # Funnel stage, not the active/removed flag.
        assert "Interview 2" in subtitle
        assert "Report date: " in subtitle

        col_a = [
            str(ws.cell(row=r, column=1).value or "") for r in range(1, ws.max_row + 1)
        ]
        for banner in (
            "─── RESUME ───",
            "─── COMPETENCE SCORES ───",
            "─── BLIND SPOTS ───",
            "─── PROCESS FINDINGS ───",
            "─── RED FLAGS ───",
            "─── FINAL VERDICT ───",
        ):
            assert banner in col_a, banner

        # Resume paragraph is merged across the five columns.
        resume_row = col_a.index("─── RESUME ───") + 2
        assert ws.cell(row=resume_row, column=1).value == (
            "14 years in maintenance planning."
        )
        assert f"A{resume_row}:E{resume_row}" in {
            str(r) for r in ws.merged_cells.ranges
        }

        # Scores table leads with the competence group.
        scores_header = col_a.index("─── COMPETENCE SCORES ───") + 2
        assert [ws.cell(row=scores_header, column=c).value for c in range(1, 7)] == [
            "Group",
            "Competence",
            "Manager score",
            "AI score",
            "Citations from transcript",
            "AI reasoning",
        ]
        assert ws.cell(row=scores_header + 1, column=1).value == "PROF: Processes"
        assert ws.cell(row=scores_header + 1, column=2).value == "Tariff analysis"

    def test_detail_sheet_names_stay_unique_when_truncated(self) -> None:
        """Two long names sharing a prefix must not collide on the
        31-char sheet-title limit."""
        long_a = "Alexandrina Konstantinopolskaya-Petrova"
        long_b = "Alexandrina Konstantinopolskaya-Ivanova"
        wb = load_workbook(
            io.BytesIO(
                render_report_xlsx(
                    vacancy={"title": "x"},
                    summary_rows=[],
                    matrix={"candidates": [], "competences": []},
                    details=[
                        {"candidate_name": long_a, "scores": []},
                        {"candidate_name": long_b, "scores": []},
                    ],
                    incomplete=[],
                    sections=["detailed_analysis"],
                )
            )
        )
        assert len(wb.sheetnames) == 2
        assert len(set(wb.sheetnames)) == 2
        assert all(len(n) <= 31 for n in wb.sheetnames)

    def test_matrix_sheet_groups_targets_divergence_and_totals(self) -> None:
        """HRP-524 — grouped rows merged in the Group column, abbreviated
        Target, ⚠ on diverging cells, ✅ on full-data AI columns, and a
        Total row that renders "n/d (p%)" with threshold colouring."""
        anna, bob = uuid.uuid4(), uuid.uuid4()
        matrix = {
            "vacancy_title": "Senior Backend Python",
            "max_score": 2.0,
            "data_disclaimer": "Data sources: Manager column is the mean…",
            "scale_disclaimer": "Scale: 1-2 (max 2.0 per competence).",
            "candidates": [
                {
                    "id": anna,
                    "name": "Anna",
                    "ai_full_data": True,
                    "manager_denominator": 4.0,
                    "ai_denominator": 4.0,
                    "manager_percent": 100.0,
                    "ai_percent": 75.0,
                },
                {
                    "id": bob,
                    "name": "Bob",
                    "ai_full_data": False,
                    "manager_denominator": 4.0,
                    "ai_denominator": 2.0,
                    "manager_percent": 25.0,
                    "ai_percent": None,
                },
            ],
            "competences": [
                {
                    "id": uuid.uuid4(),
                    "name": "Tariff analysis",
                    "group": "PROF: Processes",
                    "criticality": "critical",
                    "cells": [
                        {
                            "candidate_id": anna,
                            "manager_score": 2.0,
                            "ai_score": 2.0,
                            "ai_status": "ready",
                            "divergence": False,
                        },
                        {
                            "candidate_id": bob,
                            "manager_score": 1.0,
                            "ai_score": None,
                            "ai_status": "not_covered",
                            "divergence": False,
                        },
                    ],
                },
                {
                    "id": uuid.uuid4(),
                    "name": "Payment systems",
                    "group": "PROF: Processes",
                    "criticality": "important",
                    "cells": [
                        {
                            "candidate_id": anna,
                            "manager_score": 2.0,
                            "ai_score": 1.0,
                            "ai_status": "ready",
                            "divergence": True,
                        },
                        {
                            "candidate_id": bob,
                            "manager_score": None,
                            "ai_score": None,
                            "ai_status": "missing",
                            "divergence": False,
                        },
                    ],
                },
                {
                    "id": uuid.uuid4(),
                    "name": "Social intelligence",
                    "group": "LEAD: People",
                    "criticality": "desirable",
                    "cells": [],
                },
            ],
        }
        wb = load_workbook(
            io.BytesIO(
                render_report_xlsx(
                    vacancy=self._vacancy_payload(),
                    summary_rows=[],
                    matrix=matrix,
                    details=[],
                    incomplete=[],
                    sections=["competency_matrix"],
                )
            )
        )
        ws = wb["Matrix"]
        assert ws.cell(row=2, column=1).value == (
            "Competence matrix — 2 candidates | Senior Backend Python"
        )
        assert "Data sources" in str(ws.cell(row=3, column=1).value)
        assert "Scale:" in str(ws.cell(row=4, column=1).value)

        header = [ws.cell(row=6, column=c).value for c in range(1, 8)]
        assert header[:3] == ["Group", "Competence", "Target"]
        assert header[3:5] == ["Mgr Anna", "Mgr Bob"]
        # ✅ only on the candidate whose AI run saw a transcript.
        assert header[5] == "AI Anna ✅"
        assert header[6] == "AI Bob"

        # Target column uses the abbreviated criticality.
        assert [ws.cell(row=r, column=3).value for r in (7, 8, 9)] == [
            "Crit",
            "Imp",
            "Des",
        ]

        # The two "PROF: Processes" rows are merged in the Group column.
        merged = {str(r) for r in ws.merged_cells.ranges}
        assert "A7:A8" in merged

        # Divergence marks the cell and tints it amber. The value stays
        # numeric (HRP-524 REDO) — the ⚠ lives in the number format, so the
        # Total row's =SUM() still counts the cell.
        assert ws.cell(row=8, column=4).value == 2.0
        assert ws.cell(row=8, column=4).number_format == '0.0" ⚠"'
        assert ws.cell(row=8, column=6).value == 1.0
        assert ws.cell(row=8, column=6).number_format == '0.0" ⚠"'
        assert ws.cell(row=8, column=6).fill.start_color.rgb.endswith("FFF3CD")
        # A cell that agrees is plain — no marker in the format.
        assert ws.cell(row=7, column=4).value == 2.0
        assert "⚠" not in ws.cell(row=7, column=4).number_format
        # No score renders as an em dash, AI not-covered as n/a.
        assert ws.cell(row=8, column=5).value == "—"
        assert ws.cell(row=7, column=7).value == "n/a"

        total_row = ws.max_row
        assert str(ws.cell(row=total_row, column=2).value).startswith("TOTAL (max ")
        # Live SUM formula that displays as "<sum>/<denominator> (<pct>%)".
        anna_total = str(ws.cell(row=total_row, column=4).value)
        assert anna_total.startswith("=SUM(")
        assert '&"/4 ("' in anna_total
        # Threshold colours: Anna manager 100% emerald, Bob manager 25% rose,
        # Anna AI 75% amber (>75 is emerald, 50-75 inclusive is amber).
        assert ws.cell(row=total_row, column=4).fill.start_color.rgb.endswith("DCFCE7")
        assert ws.cell(row=total_row, column=5).fill.start_color.rgb.endswith("FECACA")
        assert ws.cell(row=total_row, column=6).fill.start_color.rgb.endswith("FFF3CD")

    def test_summary_sheet_header_disclaimer_and_notes(self) -> None:
        """HRP-523 — title, position/date line, an auto-generated
        completeness disclaimer, the six spec columns and a NOTES block
        holding the asterisk footnotes."""
        rows = [
            {
                "candidate_name": "Anna Smirnova",
                "profile_summary": "14 years in maintenance planning",
                "manager_score_text": "44/44 (100%)",
                "ai_score_text": "41/44 (93%)*",
                "data_readiness": "Full data",
                "recommendation": "Recommended",
                "note": "AI 41/44 — diverges from the manager on 3 competence(s).",
            },
            {
                "candidate_name": "Bob Petrov",
                "profile_summary": "Underground mining",
                "manager_score_text": "39/44 (89%)",
                "ai_score_text": "— (no transcript)",
                "data_readiness": "No transcript",
                "recommendation": "Additional check",
                "note": None,
            },
        ]
        wb = load_workbook(
            io.BytesIO(
                render_report_xlsx(
                    vacancy=self._vacancy_payload(),
                    summary_rows=rows,
                    matrix={"candidates": [], "competences": []},
                    details=[],
                    incomplete=[],
                    branding={"tenant_name": "Acme"},
                    sections=["summary_ranking"],
                )
            )
        )
        ws = wb["Summary"]
        assert ws.cell(row=2, column=1).value == (
            "Candidate ranking — Senior Backend Engineer | Acme"
        )
        subtitle = str(ws.cell(row=3, column=1).value)
        assert "Position: Engineering" in subtitle
        assert "Platform" in subtitle
        assert "Report date: " in subtitle

        # Mixed readiness must produce the warning variant.
        disclaimer = str(ws.cell(row=4, column=1).value)
        assert disclaimer.startswith("⚠")
        assert "1 of 2 candidates have full data" in disclaimer
        assert "1 have no interview transcript" in disclaimer

        headers = [ws.cell(row=6, column=c).value for c in range(1, 7)]
        assert headers == [
            "Candidate",
            "Profile",
            "Manager score",
            "AI score",
            "AI data",
            "Recommendation",
        ]
        assert ws.cell(row=7, column=1).value == "Anna Smirnova"
        assert ws.cell(row=7, column=4).value == "41/44 (93%)*"
        assert ws.cell(row=7, column=5).value == "Full data"
        assert ws.cell(row=8, column=4).value == "— (no transcript)"

        flat = [
            str(ws.cell(row=r, column=1).value or "") for r in range(1, ws.max_row + 1)
        ]
        assert "NOTES" in flat
        # Only the row that carries a note is footnoted.
        assert any(f.startswith("* Anna Smirnova:") for f in flat)
        assert not any(f.startswith("* Bob Petrov:") for f in flat)

    def test_summary_disclaimer_when_every_candidate_has_full_data(self) -> None:
        wb = load_workbook(
            io.BytesIO(
                render_report_xlsx(
                    vacancy={"title": "x"},
                    summary_rows=[
                        {"candidate_name": "A", "data_readiness": "Full data"}
                    ],
                    matrix={"candidates": [], "competences": []},
                    details=[],
                    incomplete=[],
                    sections=["summary_ranking"],
                )
            )
        )
        disclaimer = str(wb["Summary"].cell(row=4, column=1).value)
        assert "All candidates have resume and transcript" in disclaimer

    def test_logo_sits_top_left_on_every_sheet(self) -> None:
        """HRP-522 — 120x40 px anchored at A1 on each sheet."""
        # 1x1 transparent PNG; openpyxl only needs a decodable image.
        png = base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42m"
            "P8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
        )
        xlsx_bytes = render_report_xlsx(
            vacancy={"title": "x"},
            summary_rows=[{"candidate_name": "John Doe"}],
            matrix={"candidates": [], "competences": []},
            details=[{"candidate_name": "John Doe", "scores": []}],
            incomplete=[],
            branding={"tenant_name": "Acme", "logo_png": png},
        )
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.sheetnames) == 4
        # openpyxl reports the *intrinsic* size on reload, so the render
        # size is asserted through the anchor extent (EMU = px * 9525).
        for name in wb.sheetnames:
            ws = wb[name]
            assert len(ws._images) == 1, name
            img = ws._images[0]
            assert img.anchor._from.col == 0, name
            assert img.anchor._from.row == 0, name
            assert img.anchor.ext.cx == 120 * 9525, name
            assert img.anchor.ext.cy == 40 * 9525, name
            # Row 1 is reserved so the logo never covers the title.
            assert ws.row_dimensions[1].height == 32, name

    def test_missing_logo_does_not_break_the_export(self) -> None:
        """A tenant without a logo still gets a complete workbook."""
        wb = load_workbook(
            io.BytesIO(
                render_report_xlsx(
                    vacancy={"title": "x"},
                    summary_rows=[],
                    matrix={"candidates": [], "competences": []},
                    details=[],
                    incomplete=[],
                    branding={"tenant_name": "Acme"},
                )
            )
        )
        assert wb.sheetnames == ["Summary", "Matrix", "Incomplete data"]
        assert all(len(wb[n]._images) == 0 for n in wb.sheetnames)

    def test_sections_pick_which_sheets_are_emitted(self) -> None:
        """HRP-521 — the Generate-report dialog owns the sheet set."""
        details = [{"candidate_name": "John Doe", "scores": []}]

        only_summary = load_workbook(
            io.BytesIO(
                render_report_xlsx(
                    vacancy={"title": "x"},
                    summary_rows=[{"candidate_name": "John Doe"}],
                    matrix={"candidates": [], "competences": []},
                    details=details,
                    incomplete=[],
                    sections=["summary_ranking"],
                )
            )
        )
        assert only_summary.sheetnames == ["Summary"]

        detail_and_incomplete = load_workbook(
            io.BytesIO(
                render_report_xlsx(
                    vacancy={"title": "x"},
                    summary_rows=[{"candidate_name": "John Doe"}],
                    matrix={"candidates": [], "competences": []},
                    details=details,
                    incomplete=[],
                    sections=["detailed_analysis", "incomplete_data"],
                )
            )
        )
        assert detail_and_incomplete.sheetnames == [
            "Detail · John Doe",
            "Incomplete data",
        ]

    def test_unknown_sections_fall_back_to_the_full_workbook(self) -> None:
        """Exports created before HRP-521 carry legacy section codes —
        they must still render rather than produce an empty workbook."""
        wb = load_workbook(
            io.BytesIO(
                render_report_xlsx(
                    vacancy={"title": "x"},
                    summary_rows=[],
                    matrix={"candidates": [], "competences": []},
                    details=[],
                    incomplete=[],
                    sections=["vacancy_summary", "comparison_grid"],
                )
            )
        )
        assert wb.sheetnames == ["Summary", "Matrix", "Incomplete data"]


# ---------------------------------------------------------------------------
# Celery task end-to-end (with mocked S3)
# ---------------------------------------------------------------------------


class TestGenerateReportTask:
    async def test_resume_sourced_candidate_is_named_not_unknown(
        self, db: AsyncSession, tenant, user, monkeypatch
    ) -> None:
        """HRP-525 — a candidate parsed from a resume has no Person row,
        only ``Candidate.full_name``. The report used to key the display
        name off the Person map and rendered "Detail · Unknown"."""
        from app.config import settings as app_settings
        from app.modules.recruitment.models import Candidate

        from tests.conftest import TEST_DB_URL

        monkeypatch.setattr(app_settings, "database_url", TEST_DB_URL)

        from app.modules.recruitment.tasks import generate_report_task

        vac = await _make_vacancy(db, tenant, user, title="ResumeSourced")
        # Straight to the model: the service-level create always links a
        # Person, which is exactly the path that masked this bug.
        cand = Candidate(tenant_id=tenant.id, full_name="Nadezhda Voronova")
        db.add(cand)
        await db.flush()
        await service.attach_candidate(
            db,
            tenant.id,
            user.id,
            CandidateVacancyCreate(
                candidate_id=cand.id, vacancy_id=uuid.UUID(str(vac["id"]))
            ),
        )
        with patch(
            "app.modules.recruitment.tasks.generate_report_task.delay"
        ) as mock_delay:
            mock_delay.return_value.id = "tid"
            res = await service.enqueue_report(
                db,
                tenant.id,
                user.id,
                uuid.UUID(str(vac["id"])),
                ReportGenerateRequest(sections=["detailed_analysis"]),
            )
        await db.commit()

        captured: dict[str, bytes] = {}

        def _fake_upload(data: bytes, path: str, content_type: str) -> str:
            captured["bytes"] = data
            return f"http://example/{path}"

        monkeypatch.setattr("app.core.s3.upload_file", _fake_upload, raising=True)
        monkeypatch.setattr("app.core.s3.get_s3_client", lambda: None, raising=True)

        result = generate_report_task.run(str(res["export_id"]), str(tenant.id))
        assert result["status"] == "completed", result

        wb = load_workbook(io.BytesIO(captured["bytes"]))
        assert wb.sheetnames == ["Detail · Nadezhda Voronova"]
        ws = wb["Detail · Nadezhda Voronova"]
        assert ws.cell(row=2, column=1).value == (
            "Detailed analysis — Nadezhda Voronova"
        )

    async def test_omitted_whitelist_excludes_terminal_candidates(
        self, db: AsyncSession, tenant, user, monkeypatch
    ) -> None:
        """The wizard's "All active" scope sends no whitelist; the task
        must then leave hired/rejected/withdrawn candidates out instead
        of reporting on the whole vacancy."""
        from app.config import settings as app_settings
        from app.modules.recruitment.models import Candidate, CandidateVacancy

        from tests.conftest import TEST_DB_URL

        monkeypatch.setattr(app_settings, "database_url", TEST_DB_URL)

        from app.modules.recruitment.tasks import generate_report_task

        vac = await _make_vacancy(db, tenant, user, title="ActiveScope")
        cv_ids: dict[str, uuid.UUID] = {}
        for name in ("Active Person", "Rejected Person"):
            cand = Candidate(tenant_id=tenant.id, full_name=name)
            db.add(cand)
            await db.flush()
            link = await service.attach_candidate(
                db,
                tenant.id,
                user.id,
                CandidateVacancyCreate(
                    candidate_id=cand.id, vacancy_id=uuid.UUID(str(vac["id"]))
                ),
            )
            cv_ids[name] = uuid.UUID(str(link["id"]))
        rejected = await db.get(CandidateVacancy, cv_ids["Rejected Person"])
        assert rejected is not None
        rejected.status = "rejected"
        with patch(
            "app.modules.recruitment.tasks.generate_report_task.delay"
        ) as mock_delay:
            mock_delay.return_value.id = "tid"
            res = await service.enqueue_report(
                db,
                tenant.id,
                user.id,
                uuid.UUID(str(vac["id"])),
                ReportGenerateRequest(sections=["detailed_analysis"]),
            )
        await db.commit()

        captured: dict[str, bytes] = {}

        def _fake_upload(data: bytes, path: str, content_type: str) -> str:
            captured["bytes"] = data
            return f"http://example/{path}"

        monkeypatch.setattr("app.core.s3.upload_file", _fake_upload, raising=True)
        monkeypatch.setattr("app.core.s3.get_s3_client", lambda: None, raising=True)

        result = generate_report_task.run(str(res["export_id"]), str(tenant.id))
        assert result["status"] == "completed", result

        wb = load_workbook(io.BytesIO(captured["bytes"]))
        assert wb.sheetnames == ["Detail · Active Person"]

    async def test_partial_manager_coverage_is_an_extra_check_not_a_rejection(
        self, db: AsyncSession, tenant, user, monkeypatch
    ) -> None:
        """HRP-523 REDO — the recommendation used to divide the manager's
        score by the full competence list, so a candidate rated at the top
        of the scale on 1 of 4 competences read as 25% and was printed as
        "Not recommended". Missing evidence is an unfinished sheet, not a
        weak candidate: it belongs in "Additional check"."""
        from app.config import settings as app_settings
        from app.modules.recruitment.models import HumanAssessment, VacancyProfile

        from tests.conftest import TEST_DB_URL

        monkeypatch.setattr(app_settings, "database_url", TEST_DB_URL)

        from app.modules.recruitment.tasks import generate_report_task

        vac = await _make_vacancy(db, tenant, user, title="PartialScore")
        vacancy_id = uuid.UUID(str(vac["id"]))
        comp_ids = [uuid.uuid4() for _ in range(4)]
        db.add(
            VacancyProfile(
                tenant_id=tenant.id,
                vacancy_id=vacancy_id,
                profile_data={
                    "competences": [
                        {"id": str(cid), "name": f"Competence {n}"}
                        for n, cid in enumerate(comp_ids, start=1)
                    ]
                },
            )
        )
        cand = await _make_candidate(db, tenant, user)
        cv = await service.attach_candidate(
            db,
            tenant.id,
            user.id,
            CandidateVacancyCreate(
                candidate_id=uuid.UUID(str(cand["id"])), vacancy_id=vacancy_id
            ),
        )
        # Top of the default 5-point scale — on one competence out of four.
        db.add(
            HumanAssessment(
                tenant_id=tenant.id,
                candidate_vacancy_id=uuid.UUID(str(cv["id"])),
                competence_id=comp_ids[0],
                evaluator_id=user.id,
                score=5.0,
            )
        )
        await db.flush()

        with patch(
            "app.modules.recruitment.tasks.generate_report_task.delay"
        ) as mock_delay:
            mock_delay.return_value.id = "tid"
            res = await service.enqueue_report(
                db,
                tenant.id,
                user.id,
                vacancy_id,
                ReportGenerateRequest(sections=["summary_ranking"]),
            )
        await db.commit()

        captured: dict[str, bytes] = {}

        def _fake_upload(data: bytes, path: str, content_type: str) -> str:
            captured["bytes"] = data
            return f"http://example/{path}"

        monkeypatch.setattr("app.core.s3.upload_file", _fake_upload, raising=True)
        monkeypatch.setattr("app.core.s3.get_s3_client", lambda: None, raising=True)

        result = generate_report_task.run(str(res["export_id"]), str(tenant.id))
        assert result["status"] == "completed", result

        ws = load_workbook(io.BytesIO(captured["bytes"]))["Summary"]
        assert ws.cell(row=6, column=6).value == "Recommendation"
        assert ws.cell(row=7, column=6).value == "Additional check"
        # The displayed score keeps the full denominator — coverage is
        # visible to the reader even though it no longer drives the verdict.
        assert ws.cell(row=7, column=3).value == "5.0/20.0 (25.0%)"

    async def test_task_writes_xlsx_and_marks_completed(
        self, db: AsyncSession, tenant, user, monkeypatch
    ) -> None:
        # The Celery task spins up its own *sync* SQLAlchemy engine via
        # ``settings.database_url``. In tests we run against
        # ``hrpulsar_test`` (see tests/conftest.py), so we point the task
        # at the same DB before triggering it; otherwise the export row
        # written via ``enqueue_report`` is invisible to the worker.
        from app.config import settings as app_settings

        from tests.conftest import TEST_DB_URL

        monkeypatch.setattr(app_settings, "database_url", TEST_DB_URL)

        from app.modules.recruitment.tasks import generate_report_task

        vac = await _make_vacancy(db, tenant, user, title="ReportVac")
        cand = await _make_candidate(db, tenant, user)
        await service.attach_candidate(
            db,
            tenant.id,
            user.id,
            CandidateVacancyCreate(
                candidate_id=uuid.UUID(str(cand["id"])),
                vacancy_id=uuid.UUID(str(vac["id"])),
            ),
        )
        with patch(
            "app.modules.recruitment.tasks.generate_report_task.delay"
        ) as mock_delay:
            mock_delay.return_value.id = "tid"
            res = await service.enqueue_report(
                db,
                tenant.id,
                user.id,
                uuid.UUID(str(vac["id"])),
                ReportGenerateRequest(
                    sections=["summary_ranking", "competency_matrix"],
                ),
            )
        export_id = res["export_id"]
        # Force the test session to flush so the sync engine sees the row.
        await db.commit()

        captured: dict[str, bytes] = {}

        def _fake_upload(data: bytes, path: str, content_type: str) -> str:
            captured["bytes"] = data
            captured["path"] = path
            return f"http://example/{path}"

        # Patch the symbol the task imports lazily.
        monkeypatch.setattr("app.core.s3.upload_file", _fake_upload, raising=True)
        monkeypatch.setattr("app.core.s3.get_s3_client", lambda: None, raising=True)

        # Run the task synchronously (bypass the broker).
        result = generate_report_task.run(str(export_id), str(tenant.id))
        assert result["status"] == "completed", result

        # Re-read via the async session (force a fresh round-trip so we do
        # not see a cached pre-task object).
        await db.commit()
        await db.close()
        export = (
            await db.execute(
                select(ConsolidatedReport).where(ConsolidatedReport.id == export_id)
            )
        ).scalar_one()
        assert export.status == "completed"
        assert export.file_id is not None
        assert export.completed_at is not None

        # The bytes must be a real XLSX (parsable by openpyxl). HRP-521:
        # the export requested Summary + Matrix only, so those are the
        # only sheets the workbook may carry — the sheet set now comes
        # from the Generate-report dialog rather than being fixed.
        wb = load_workbook(io.BytesIO(captured["bytes"]))
        assert wb.sheetnames == ["Summary", "Matrix"]

    async def test_task_marks_failed_when_upload_returns_none(
        self, db: AsyncSession, tenant, user, monkeypatch
    ) -> None:
        from app.config import settings as app_settings

        from tests.conftest import TEST_DB_URL

        monkeypatch.setattr(app_settings, "database_url", TEST_DB_URL)

        from app.modules.recruitment.tasks import generate_report_task

        vac = await _make_vacancy(db, tenant, user)
        with patch(
            "app.modules.recruitment.tasks.generate_report_task.delay"
        ) as mock_delay:
            mock_delay.return_value.id = "tid"
            res = await service.enqueue_report(
                db,
                tenant.id,
                user.id,
                uuid.UUID(str(vac["id"])),
                ReportGenerateRequest(),
            )
        await db.commit()

        monkeypatch.setattr(
            "app.core.s3.upload_file", lambda *a, **k: None, raising=True
        )
        monkeypatch.setattr("app.core.s3.get_s3_client", lambda: None, raising=True)

        result = generate_report_task.run(str(res["export_id"]), str(tenant.id))
        assert result["status"] == "failed", result

        await db.commit()
        await db.close()
        export = (
            await db.execute(
                select(ConsolidatedReport).where(
                    ConsolidatedReport.id == res["export_id"]
                )
            )
        ).scalar_one()
        assert export.status == "failed"
        assert "storage" in (export.error or "").lower()

    async def test_task_marks_failed_when_requesting_user_is_gone(
        self, db: AsyncSession, tenant, user, monkeypatch
    ) -> None:
        # ConsolidatedReport.generated_by has ON DELETE SET NULL so a
        # deleted user leaves the export with NULL. files.uploaded_by is
        # FK users.id NOT NULL — the worker must fail-fast with a clear
        # error instead of either crashing inside the FK check or writing
        # a wrong UUID into uploaded_by (regression for a bug found in
        # review).
        from app.config import settings as app_settings

        from tests.conftest import TEST_DB_URL

        monkeypatch.setattr(app_settings, "database_url", TEST_DB_URL)

        from app.modules.recruitment.tasks import generate_report_task

        vac = await _make_vacancy(db, tenant, user)
        with patch(
            "app.modules.recruitment.tasks.generate_report_task.delay"
        ) as mock_delay:
            mock_delay.return_value.id = "tid"
            res = await service.enqueue_report(
                db,
                tenant.id,
                user.id,
                uuid.UUID(str(vac["id"])),
                ReportGenerateRequest(sections=["summary_ranking"]),
            )
        # Simulate the user being deleted between enqueue and worker run.
        export = (
            await db.execute(
                select(ConsolidatedReport).where(
                    ConsolidatedReport.id == res["export_id"]
                )
            )
        ).scalar_one()
        export.generated_by = None
        await db.commit()
        await db.close()

        # Upload should never be reached — fail-fast happens before it.
        called: dict[str, bool] = {}

        def _should_not_be_called(*_args, **_kwargs):
            called["upload"] = True
            return "http://example/x"

        monkeypatch.setattr(
            "app.core.s3.upload_file", _should_not_be_called, raising=True
        )

        result = generate_report_task.run(str(res["export_id"]), str(tenant.id))
        assert result["status"] == "failed"
        assert "user" in result["error"].lower()
        assert "upload" not in called

        await db.commit()
        export_after = (
            await db.execute(
                select(ConsolidatedReport).where(
                    ConsolidatedReport.id == res["export_id"]
                )
            )
        ).scalar_one()
        assert export_after.status == "failed"
        assert export_after.file_id is None


# ---------------------------------------------------------------------------
# R4d — Inline XLSX preview (FR-23 / SCR-83)
# ---------------------------------------------------------------------------


class TestReportPreview:
    """``GET /recruitment/reports/{id}/preview`` returns every sheet as JSON."""

    @staticmethod
    async def _seed_completed_report(db, tenant, user):
        from app.modules.recruitment.models import ConsolidatedReport
        from app.modules.storage.models import File

        vac = await _make_vacancy(db, tenant, user)
        file_row = File(
            tenant_id=tenant.id,
            name="report.xlsx",
            original_name="report.xlsx",
            path=f"{tenant.id}/reports/{uuid.uuid4()}.xlsx",
            size=1024,
            mime_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            uploaded_by=user.id,
            entity_type="report",
            entity_id=uuid.uuid4(),
        )
        db.add(file_row)
        await db.flush()
        export = ConsolidatedReport(
            tenant_id=tenant.id,
            vacancy_id=uuid.UUID(str(vac["id"])),
            sections=["summary_ranking"],
            status="completed",
            generated_by=user.id,
            file_id=file_row.id,
        )
        db.add(export)
        await db.commit()
        await db.refresh(export)
        await db.refresh(file_row)
        return export, file_row

    async def test_preview_returns_sheets_when_xlsx_is_readable(
        self, db: AsyncSession, tenant, user, monkeypatch
    ):
        from openpyxl import Workbook

        export, file_row = await self._seed_completed_report(db, tenant, user)

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Header", "Value"])
        ws.append(["Vacancy", "Senior Python"])
        ws.append(["Candidates", 3])
        wb.create_sheet("Notes").append(["x"])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        class _Body:
            def __init__(self, data: bytes) -> None:
                self._data = data

            def read(self) -> bytes:
                return self._data

        class _Client:
            def get_object(self, **kwargs):
                return {"Body": _Body(xlsx_bytes)}

        monkeypatch.setattr("app.core.s3.get_s3_client", lambda: _Client())
        # Force a Redis miss + skip cache write so the test does not
        # depend on a running Redis.
        monkeypatch.setattr(
            "redis.asyncio.from_url",
            lambda *a, **kw: (_ for _ in ()).throw(RuntimeError("no redis")),
        )

        result = await service.get_report_preview(db, tenant.id, export.id)
        assert result["export_id"] == str(export.id)
        names = [s["name"] for s in result["sheets"]]
        assert "Summary" in names and "Notes" in names
        summary = next(s for s in result["sheets"] if s["name"] == "Summary")
        assert summary["cells"][0] == ["Header", "Value"]
        assert summary["cells"][1] == ["Vacancy", "Senior Python"]
        assert summary["cells"][2] == ["Candidates", 3]
        assert result["truncated"] is False

    async def test_preview_409_when_report_not_completed(
        self, db: AsyncSession, tenant, user
    ):
        from app.modules.recruitment.models import ConsolidatedReport

        vac = await _make_vacancy(db, tenant, user)
        export = ConsolidatedReport(
            tenant_id=tenant.id,
            vacancy_id=uuid.UUID(str(vac["id"])),
            sections=["summary_ranking"],
            status="processing",
            generated_by=user.id,
        )
        db.add(export)
        await db.commit()
        await db.refresh(export)

        with pytest.raises(HTTPException) as exc:
            await service.get_report_preview(db, tenant.id, export.id)
        assert exc.value.status_code == 409

    async def test_preview_404_when_export_missing(self, db: AsyncSession, tenant):
        with pytest.raises(HTTPException) as exc:
            await service.get_report_preview(db, tenant.id, uuid.uuid4())
        assert exc.value.status_code == 404
