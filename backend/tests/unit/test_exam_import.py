"""GF9: Exam Excel Import — import questions from Excel with validation."""

import uuid
from io import BytesIO

import pytest
from app.modules.exam.models import ExamQuestion, QuestionOption
from app.modules.exam.service import (
    create_mass_exam,
    generate_import_template,
    import_questions_from_excel,
)
from fastapi import HTTPException
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession


def _make_excel(*rows) -> bytes:
    """Create an Excel file with given rows (header auto-added)."""
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "question_text",
            "question_type",
            "option_1",
            "option_2",
            "option_3",
            "option_4",
            "option_5",
            "option_6",
            "correct_options",
            "weight",
        ]
    )
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class _Stub:
    def __init__(self, **kwargs):
        self.__dict__.update(kwargs)


class TestExamImport:
    async def _make_exam(self, db, tenant, user):
        data = _Stub(title="Python Basics", description=None, ended_at=None)
        return await create_mass_exam(db, tenant.id, user.id, data)

    async def test_import_single_choice(self, db: AsyncSession, tenant, user):
        mass_exam = await self._make_exam(db, tenant, user)
        excel = _make_excel(
            ["What is 2+2?", "single_choice", "3", "4", "5", "", "", "", "2", "1"],
        )
        result = await import_questions_from_excel(
            db, tenant.id, mass_exam["id"], excel
        )

        assert result["imported"] == 1
        assert result["errors"] == []

        q_result = await db.execute(
            select(ExamQuestion).where(ExamQuestion.mass_exam_id == mass_exam["id"])
        )
        q = q_result.scalar_one()
        assert q.title == "What is 2+2?"
        assert q.question_type == "single_choice"
        assert q.weight == 1

        opts = await db.execute(
            select(QuestionOption)
            .where(QuestionOption.question_id == q.id)
            .order_by(QuestionOption.sort_index)
        )
        options = opts.scalars().all()
        assert len(options) == 3
        assert options[0].title == "3"
        assert not options[0].is_correct
        assert options[1].title == "4"
        assert options[1].is_correct

    async def test_import_multiple_choice(self, db: AsyncSession, tenant, user):
        mass_exam = await self._make_exam(db, tenant, user)
        excel = _make_excel(
            [
                "Select primes",
                "multiple_choice",
                "2",
                "3",
                "4",
                "5",
                "",
                "",
                "1,2,4",
                "2",
            ],
        )
        result = await import_questions_from_excel(
            db, tenant.id, mass_exam["id"], excel
        )

        assert result["imported"] == 1

        q_result = await db.execute(
            select(ExamQuestion).where(ExamQuestion.mass_exam_id == mass_exam["id"])
        )
        q = q_result.scalar_one()
        assert q.question_type == "multiple_choice"
        assert q.weight == 2

        opts = await db.execute(
            select(QuestionOption)
            .where(QuestionOption.question_id == q.id)
            .order_by(QuestionOption.sort_index)
        )
        options = opts.scalars().all()
        assert len(options) == 4
        assert options[0].is_correct  # index 1
        assert options[1].is_correct  # index 2
        assert not options[2].is_correct  # index 3
        assert options[3].is_correct  # index 4

    async def test_import_essay(self, db: AsyncSession, tenant, user):
        mass_exam = await self._make_exam(db, tenant, user)
        excel = _make_excel(
            ["Explain polymorphism", "essay", "", "", "", "", "", "", "", "3"],
        )
        result = await import_questions_from_excel(
            db, tenant.id, mass_exam["id"], excel
        )

        assert result["imported"] == 1
        q_result = await db.execute(
            select(ExamQuestion).where(ExamQuestion.mass_exam_id == mass_exam["id"])
        )
        q = q_result.scalar_one()
        assert q.question_type == "essay"
        assert q.weight == 3

    async def test_import_duplicate_detection(self, db: AsyncSession, tenant, user):
        mass_exam = await self._make_exam(db, tenant, user)
        excel = _make_excel(
            ["What is 2+2?", "single_choice", "3", "4", "", "", "", "", "2", "1"],
            ["What is 2+2?", "single_choice", "5", "6", "", "", "", "", "1", "1"],
        )
        result = await import_questions_from_excel(
            db, tenant.id, mass_exam["id"], excel
        )

        assert result["imported"] == 1
        assert len(result["errors"]) == 1
        assert "Duplicate" in result["errors"][0]["error"]

    async def test_import_invalid_type(self, db: AsyncSession, tenant, user):
        mass_exam = await self._make_exam(db, tenant, user)
        excel = _make_excel(
            ["Question", "invalid_type", "A", "B", "", "", "", "", "1", "1"],
        )
        result = await import_questions_from_excel(
            db, tenant.id, mass_exam["id"], excel
        )

        assert result["imported"] == 0
        assert len(result["errors"]) == 1
        assert "Invalid question type" in result["errors"][0]["error"]

    async def test_import_choice_needs_options(self, db: AsyncSession, tenant, user):
        mass_exam = await self._make_exam(db, tenant, user)
        excel = _make_excel(
            ["Question?", "single_choice", "Only one", "", "", "", "", "", "1", "1"],
        )
        result = await import_questions_from_excel(
            db, tenant.id, mass_exam["id"], excel
        )

        assert result["imported"] == 0
        assert len(result["errors"]) == 1
        assert "at least 2 options" in result["errors"][0]["error"]

    async def test_import_multiple_rows(self, db: AsyncSession, tenant, user):
        mass_exam = await self._make_exam(db, tenant, user)
        excel = _make_excel(
            ["Q1", "single_choice", "A", "B", "", "", "", "", "1", "1"],
            ["Q2", "single_choice", "C", "D", "", "", "", "", "2", "2"],
            ["Q3", "essay", "", "", "", "", "", "", "", "5"],
        )
        result = await import_questions_from_excel(
            db, tenant.id, mass_exam["id"], excel
        )

        assert result["imported"] == 3
        assert result["total_rows"] == 3
        assert result["errors"] == []

    async def test_import_mass_exam_not_found(self, db: AsyncSession, tenant, user):
        excel = _make_excel(["Q", "essay", "", "", "", "", "", "", "", "1"])
        with pytest.raises(HTTPException):
            await import_questions_from_excel(db, tenant.id, uuid.uuid4(), excel)

    async def test_generate_template(self, db: AsyncSession):
        data = generate_import_template()
        assert len(data) > 0
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(data))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][0] == "question_text"
        assert len(rows) == 4  # header + 3 examples
