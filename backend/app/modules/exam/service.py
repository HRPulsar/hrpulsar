import uuid
from datetime import datetime, timezone
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select, update
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.s3 import get_presigned_url
from app.modules.employee.models import Employee
from app.modules.exam.models import (
    Exam,
    ExamAnswer,
    ExamPassMark,
    ExamQuestion,
    MassExam,
    QuestionOption,
)
from app.modules.exam.schemas import ALLOWED_STATUS_TRANSITIONS, TERMINAL_STATUSES
from app.modules.storage.models import File

# --- MassExam ---


async def create_mass_exam(
    db: AsyncSession, tenant_id: uuid.UUID, initiator_id: uuid.UUID, data
) -> dict:

    me = MassExam(
        tenant_id=tenant_id,
        title=data.title,
        description=data.description,
        initiator_id=initiator_id,
        ended_at=data.ended_at,
    )
    db.add(me)
    await db.commit()
    await db.refresh(me)
    return _me_to_read(me)


async def list_mass_exams(db: AsyncSession, tenant_id: uuid.UUID) -> list[dict]:
    result = await db.execute(
        select(MassExam)
        .where(MassExam.tenant_id == tenant_id)
        .order_by(MassExam.created_at.desc())
    )
    return [_me_to_read(m) for m in result.scalars().all()]


async def get_mass_exam_detail(
    db: AsyncSession, tenant_id: uuid.UUID, mass_exam_id: uuid.UUID
) -> dict:
    result = await db.execute(
        select(MassExam)
        .options(
            selectinload(MassExam.questions).selectinload(ExamQuestion.options),
            selectinload(MassExam.pass_marks),
        )
        .where(MassExam.id == mass_exam_id, MassExam.tenant_id == tenant_id)
    )
    me = result.scalar_one_or_none()
    if not me:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Mass exam not found")

    data = _me_to_read(me)
    questions = []
    for q in me.questions:
        img_url = await _resolve_image_url(db, q.image_file_id)
        questions.append(
            {
                "id": q.id,
                "title": q.title,
                "description": q.description,
                "question_type": q.question_type,
                "sort_index": q.sort_index,
                "weight": q.weight,
                "is_active": q.is_active,
                "image_file_id": q.image_file_id,
                "image_url": img_url,
                "options": [
                    {
                        "id": o.id,
                        "title": o.title,
                        "sort_index": o.sort_index,
                        "is_correct": o.is_correct,
                    }
                    for o in q.options
                ],
            }
        )
    data["questions"] = questions
    data["pass_marks"] = [
        {
            "id": pm.id,
            "grade_id": pm.grade_id,
            "specialization_id": pm.specialization_id,
            "min_score_percent": pm.min_score_percent,
            "min_score_points": pm.min_score_points,
        }
        for pm in me.pass_marks
    ]
    assigned_count = await db.scalar(
        select(func.count(Exam.id)).where(Exam.mass_exam_id == mass_exam_id)
    )
    data["assigned_count"] = int(assigned_count or 0)
    return data


async def update_mass_exam(
    db: AsyncSession, tenant_id: uuid.UUID, mass_exam_id: uuid.UUID, data
) -> dict:
    me = await db.get(MassExam, mass_exam_id)
    if not me or me.tenant_id != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Mass exam not found")

    if me.status in TERMINAL_STATUSES:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Cannot edit a finalized exam",
        )

    payload = data.model_dump(exclude_unset=True)
    if "title" in payload and payload["title"] is not None:
        me.title = payload["title"]
    if "description" in payload:
        me.description = payload["description"]
    if "ended_at" in payload:
        me.ended_at = payload["ended_at"]

    await db.commit()
    await db.refresh(me)
    return _me_to_read(me)


async def change_mass_exam_status(
    db: AsyncSession, tenant_id: uuid.UUID, mass_exam_id: uuid.UUID, new_status: str
) -> dict:

    # FOR UPDATE + populate_existing (same pattern as finish_exam): without
    # the lock a participant finishing concurrently can auto-flip the exam
    # to done while this transition validates against a stale in_progress —
    # a manual cancel would then land on top of done, bypassing the state
    # machine.
    me = (
        await db.execute(
            select(MassExam)
            .where(MassExam.id == mass_exam_id)
            .with_for_update()
            .execution_options(populate_existing=True)
        )
    ).scalar_one_or_none()
    if not me or me.tenant_id != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Mass exam not found")

    allowed = ALLOWED_STATUS_TRANSITIONS.get(me.status, set())
    if new_status not in allowed:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            (
                f"Cannot transition from '{me.status}' to '{new_status}'. "
                f"Allowed: {sorted(allowed) if allowed else 'none'}"
            ),
        )

    if me.status == "draft" and new_status == "sent":
        await _assert_send_prerequisites(db, me)
        # HRP-328: questions stay editable while in draft, so max_score
        # frozen at assign time can be stale. The question set locks on
        # send — recompute every participant's max_score here.
        max_score = await db.scalar(
            select(func.coalesce(func.sum(ExamQuestion.weight), 0)).where(
                ExamQuestion.mass_exam_id == me.id,
                ExamQuestion.is_active.is_(True),
            )
        )
        exams_result = await db.execute(
            select(Exam).where(Exam.mass_exam_id == me.id)
        )
        for e in exams_result.scalars().all():
            e.max_score = max_score

    me.status = new_status
    if new_status == "sent" and me.started_at is None:
        # left empty intentionally — started_at is set when first employee starts
        pass
    if new_status == "in_progress" and not me.started_at:
        me.started_at = datetime.now(timezone.utc)
    elif new_status in {"done", "cancelled"}:
        me.finished_at = datetime.now(timezone.utc)
        # HRP-236: closing the procedure cascades to participant surveys.
        # Cancel voids every survey (submitted scores stay on the row);
        # complete keeps finished surveys done and voids the unfinished.
        cascade_from = (
            ["assigned", "in_progress"]
            if new_status == "done"
            else ["assigned", "in_progress", "done"]
        )
        await db.execute(
            update(Exam)
            .where(
                Exam.mass_exam_id == me.id,
                Exam.status.in_(cascade_from),
            )
            .values(status="cancelled")
        )
    await db.commit()
    await db.refresh(me)
    return _me_to_read(me)


async def _assert_send_prerequisites(db: AsyncSession, me: MassExam) -> None:
    """draft → sent requires ≥1 question, ≥1 assigned employee, future deadline."""
    if me.ended_at is not None:
        cutoff = datetime.now(timezone.utc).date()
        ended = me.ended_at
        ended_date = (
            ended.astimezone(timezone.utc).date()
            if ended.tzinfo is not None
            else ended.replace(tzinfo=timezone.utc).date()
        )
        if ended_date < cutoff:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST, "Deadline in the past"
            )

    q_count = await db.scalar(
        select(func.count(ExamQuestion.id)).where(
            ExamQuestion.mass_exam_id == me.id, ExamQuestion.is_active.is_(True)
        )
    )
    if not q_count:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST, "Add at least one question before sending"
        )

    e_count = await db.scalar(
        select(func.count(Exam.id)).where(Exam.mass_exam_id == me.id)
    )
    if not e_count:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST, "Assign at least one employee before sending"
        )


def _me_to_read(me: MassExam) -> dict:
    return {
        "id": me.id,
        "title": me.title,
        "description": me.description,
        "initiator_id": me.initiator_id,
        "status": me.status,
        "tenant_id": me.tenant_id,
        "started_at": me.started_at,
        "ended_at": me.ended_at,
        "finished_at": me.finished_at,
        "created_at": me.created_at,
    }


# --- Questions ---


async def _resolve_image_url(db: AsyncSession, file_id: uuid.UUID | None) -> str | None:
    if not file_id:
        return None
    f = await db.get(File, file_id)
    if not f:
        return None
    return get_presigned_url(f.path)


async def add_question(
    db: AsyncSession, tenant_id: uuid.UUID, mass_exam_id: uuid.UUID, data
) -> dict:
    me = await db.get(MassExam, mass_exam_id)
    if not me or me.tenant_id != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Mass exam not found")

    q = ExamQuestion(
        mass_exam_id=mass_exam_id,
        title=data.title,
        description=data.description,
        question_type=data.question_type,
        sort_index=data.sort_index,
        weight=data.weight,
        image_file_id=getattr(data, "image_file_id", None),
    )
    db.add(q)
    await db.flush()

    for opt in data.options:
        db.add(
            QuestionOption(
                question_id=q.id,
                title=opt.title,
                sort_index=opt.sort_index,
                is_correct=opt.is_correct,
            )
        )

    await db.commit()
    await db.refresh(q)
    # Reload with options
    result = await db.execute(
        select(ExamQuestion)
        .options(selectinload(ExamQuestion.options))
        .where(ExamQuestion.id == q.id)
    )
    q = result.scalar_one()
    img_url = await _resolve_image_url(db, q.image_file_id)
    return {
        "id": q.id,
        "title": q.title,
        "description": q.description,
        "question_type": q.question_type,
        "sort_index": q.sort_index,
        "weight": q.weight,
        "is_active": q.is_active,
        "image_file_id": q.image_file_id,
        "image_url": img_url,
        "options": [
            {
                "id": o.id,
                "title": o.title,
                "sort_index": o.sort_index,
                "is_correct": o.is_correct,
            }
            for o in q.options
        ],
    }


async def update_question(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    mass_exam_id: uuid.UUID,
    question_id: uuid.UUID,
    data,
) -> dict:
    """HRP-228: edit a question while the parent exam is still in draft.

    Replacing ``options`` wipes existing rows and re-creates them — IDs are
    not preserved. The caller is the form on /exams/[uuid], which always
    submits the full options list.
    """
    me = await db.get(MassExam, mass_exam_id)
    if not me or me.tenant_id != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Mass exam not found")
    if me.status != "draft":
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Questions can only be edited while the exam is in draft",
        )

    q_result = await db.execute(
        select(ExamQuestion)
        .options(selectinload(ExamQuestion.options))
        .where(
            ExamQuestion.id == question_id,
            ExamQuestion.mass_exam_id == mass_exam_id,
        )
    )
    q = q_result.scalar_one_or_none()
    if q is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Question not found")

    payload = data.model_dump(exclude_unset=True)
    if "title" in payload and payload["title"] is not None:
        q.title = payload["title"]
    if "description" in payload:
        q.description = payload["description"]
    if "question_type" in payload and payload["question_type"] is not None:
        q.question_type = payload["question_type"]
    if "weight" in payload and payload["weight"] is not None:
        q.weight = payload["weight"]
    if "image_file_id" in payload:
        q.image_file_id = payload["image_file_id"]

    if data.options is not None:
        # Re-validate against the persisted question type now that we know it
        # (Pydantic only had ``question_type`` if explicitly supplied).
        effective_type = q.question_type
        new_options = data.options
        if effective_type != "essay":
            non_empty = [o for o in new_options if o.title.strip()]
            if len(non_empty) < 2:
                raise HTTPException(
                    status.HTTP_400_BAD_REQUEST,
                    "Choice questions require at least 2 options",
                )
            correct = [o for o in non_empty if o.is_correct]
            if not correct:
                raise HTTPException(
                    status.HTTP_400_BAD_REQUEST,
                    "Choice questions require at least one correct option",
                )
            if effective_type == "single_choice" and len(correct) > 1:
                raise HTTPException(
                    status.HTTP_400_BAD_REQUEST,
                    "Single choice questions allow only one correct option",
                )
        # Use the collection so the ``cascade="all, delete-orphan"`` rule
        # cleans up unlinked options without an extra SQL DELETE.
        q.options.clear()
        await db.flush()
        for opt in new_options:
            q.options.append(
                QuestionOption(
                    title=opt.title,
                    sort_index=opt.sort_index,
                    is_correct=opt.is_correct,
                )
            )

    await db.commit()
    result = await db.execute(
        select(ExamQuestion)
        .options(selectinload(ExamQuestion.options))
        .where(ExamQuestion.id == q.id)
    )
    q = result.scalar_one()
    img_url = await _resolve_image_url(db, q.image_file_id)
    return {
        "id": q.id,
        "title": q.title,
        "description": q.description,
        "question_type": q.question_type,
        "sort_index": q.sort_index,
        "weight": q.weight,
        "is_active": q.is_active,
        "image_file_id": q.image_file_id,
        "image_url": img_url,
        "options": [
            {
                "id": o.id,
                "title": o.title,
                "sort_index": o.sort_index,
                "is_correct": o.is_correct,
            }
            for o in q.options
        ],
    }


async def delete_question(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    mass_exam_id: uuid.UUID,
    question_id: uuid.UUID,
) -> None:
    """HRP-228: hard-delete a question (with its options) while the parent
    exam is still in draft. Cascades wipe ``question_options``.
    """
    me = await db.get(MassExam, mass_exam_id)
    if not me or me.tenant_id != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Mass exam not found")
    if me.status != "draft":
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Questions can only be deleted while the exam is in draft",
        )
    q = await db.get(ExamQuestion, question_id)
    if not q or q.mass_exam_id != mass_exam_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Question not found")
    await db.delete(q)
    await db.commit()


# --- PassMark ---


def _assert_pass_mark_editable(me: MassExam) -> None:
    # HRP-226 redo: Pass Mark CRUD is available on any active exam
    # (draft / sent / in_progress); only terminal statuses freeze it.
    if me.status in TERMINAL_STATUSES:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Pass marks cannot be edited on a finalized exam",
        )


def _pass_mark_to_read(pm: ExamPassMark) -> dict:
    return {
        "id": pm.id,
        "grade_id": pm.grade_id,
        "specialization_id": pm.specialization_id,
        "min_score_percent": pm.min_score_percent,
        "min_score_points": pm.min_score_points,
    }


async def _get_editable_pass_mark(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    mass_exam_id: uuid.UUID,
    pass_mark_id: uuid.UUID,
) -> ExamPassMark:
    """Shared update/delete preamble: exam lookup, terminal gate, row lookup."""
    me = await db.get(MassExam, mass_exam_id)
    if not me or me.tenant_id != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Mass exam not found")
    _assert_pass_mark_editable(me)
    pm = await db.get(ExamPassMark, pass_mark_id)
    if not pm or pm.mass_exam_id != mass_exam_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Pass mark not found")
    return pm


async def _validate_pass_mark_refs(
    db: AsyncSession, tenant_id: uuid.UUID, payload: dict
) -> None:
    """Reject grade / specialization ids that don't resolve to a dictionary
    item of the right type visible to the tenant — otherwise a bogus UUID
    surfaces as a bare FK IntegrityError (500) and a foreign tenant's (or
    wrong-type) id would be accepted."""
    from app.modules.dictionary.models import DictionaryItem

    expected_types = {"grade_id": "grade", "specialization_id": "specialization"}
    for field, expected_type in expected_types.items():
        ref = payload.get(field)
        if ref is None:
            continue
        item = await db.get(DictionaryItem, ref)
        if (
            not item
            or item.type != expected_type
            or (item.tenant_id is not None and item.tenant_id != tenant_id)
        ):
            raise HTTPException(
                status.HTTP_404_NOT_FOUND, f"Dictionary item for {field} not found"
            )


async def add_pass_mark(
    db: AsyncSession, tenant_id: uuid.UUID, mass_exam_id: uuid.UUID, data
) -> dict:
    me = await db.get(MassExam, mass_exam_id)
    if not me or me.tenant_id != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Mass exam not found")
    _assert_pass_mark_editable(me)

    existing = await db.scalar(
        select(func.count(ExamPassMark.id)).where(
            ExamPassMark.mass_exam_id == mass_exam_id
        )
    )
    if existing:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST, "Pass mark already exists for this exam"
        )
    if data.min_score_percent is None and data.min_score_points is None:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Pass mark requires a minimum score (percent or points)",
        )
    await _validate_pass_mark_refs(
        db,
        tenant_id,
        {"grade_id": data.grade_id, "specialization_id": data.specialization_id},
    )

    pm = ExamPassMark(
        mass_exam_id=mass_exam_id,
        grade_id=data.grade_id,
        specialization_id=data.specialization_id,
        min_score_percent=data.min_score_percent,
        min_score_points=data.min_score_points,
    )
    db.add(pm)
    await db.commit()
    await db.refresh(pm)
    return _pass_mark_to_read(pm)


async def update_pass_mark(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    mass_exam_id: uuid.UUID,
    pass_mark_id: uuid.UUID,
    data,
) -> dict:
    """HRP-226 redo: edit the existing pass mark in place."""
    pm = await _get_editable_pass_mark(db, tenant_id, mass_exam_id, pass_mark_id)
    payload = data.model_dump(exclude_unset=True)
    await _validate_pass_mark_refs(db, tenant_id, payload)
    for field, value in payload.items():
        setattr(pm, field, value)
    if pm.min_score_percent is None and pm.min_score_points is None:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Pass mark requires a minimum score (percent or points)",
        )
    await db.commit()
    await db.refresh(pm)
    return _pass_mark_to_read(pm)


async def delete_pass_mark(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    mass_exam_id: uuid.UUID,
    pass_mark_id: uuid.UUID,
) -> None:
    pm = await _get_editable_pass_mark(db, tenant_id, mass_exam_id, pass_mark_id)
    await db.delete(pm)
    await db.commit()


# --- Assign employees ---


async def assign_employees(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    mass_exam_id: uuid.UUID,
    employee_ids: list[uuid.UUID],
) -> list[dict]:
    me = await db.get(MassExam, mass_exam_id)
    if not me or me.tenant_id != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Mass exam not found")

    # Calculate max score
    q_result = await db.execute(
        select(ExamQuestion).where(
            ExamQuestion.mass_exam_id == mass_exam_id, ExamQuestion.is_active.is_(True)
        )
    )
    max_score = sum(q.weight for q in q_result.scalars().all())

    # Filter to employees that belong to this tenant; silently drop unknowns.
    valid_ids: set[uuid.UUID] = set()
    if employee_ids:
        rows = await db.execute(
            select(Employee.id).where(
                Employee.tenant_id == tenant_id, Employee.id.in_(employee_ids)
            )
        )
        valid_ids = {row[0] for row in rows.all()}

    # Skip employees already assigned to this mass exam to avoid duplicates.
    if valid_ids:
        existing_rows = await db.execute(
            select(Exam.employee_id).where(
                Exam.mass_exam_id == mass_exam_id, Exam.employee_id.in_(valid_ids)
            )
        )
        already = {row[0] for row in existing_rows.all()}
        valid_ids -= already

    exams = []
    for eid in valid_ids:
        exam = Exam(
            tenant_id=tenant_id,
            mass_exam_id=mass_exam_id,
            employee_id=eid,
            max_score=max_score,
        )
        db.add(exam)
        exams.append(exam)

    await db.commit()
    return [
        {
            "id": e.id,
            "mass_exam_id": e.mass_exam_id,
            "employee_id": e.employee_id,
            "status": e.status,
            "score": e.score,
            "max_score": e.max_score,
            "started_at": e.started_at,
            "finished_at": e.finished_at,
            "tenant_id": e.tenant_id,
            "created_at": e.created_at,
        }
        for e in exams
    ]


# --- Mass Exam Results ---


def _exam_passed(
    score: int | None,
    max_score: int | None,
    pass_mark: ExamPassMark | None,
) -> bool | None:
    """HRP-364: pass/fail against the exam's Pass Mark.

    With a configured mark every set threshold must be met (percent and/or
    points); without one the default is 60%, matching the UI hint "No pass
    mark configured (default: 60%)". ``passed`` is always derived at read
    time, so editing the mark on an active exam (HRP-226) transparently
    re-grades already-finished participants.
    """
    if score is None:
        return None
    thresholds_set = pass_mark is not None and (
        pass_mark.min_score_percent is not None
        or pass_mark.min_score_points is not None
    )
    if not thresholds_set:
        if not max_score:
            return None
        return (score / max_score * 100) >= 60
    assert pass_mark is not None
    if pass_mark.min_score_percent is not None:
        if not max_score:
            # A percent threshold is undecidable without a max score.
            return None
        if (score / max_score * 100) < pass_mark.min_score_percent:
            return False
    return not (
        pass_mark.min_score_points is not None
        and score < pass_mark.min_score_points
    )


async def _load_pass_mark(
    db: AsyncSession, mass_exam_id: uuid.UUID
) -> ExamPassMark | None:
    return (
        await db.execute(
            select(ExamPassMark).where(ExamPassMark.mass_exam_id == mass_exam_id)
        )
    ).scalars().first()


async def list_mass_exam_results(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    mass_exam_id: uuid.UUID,
    visible_employee_ids: set[uuid.UUID] | None = None,
) -> list[dict]:
    me = await db.get(MassExam, mass_exam_id)
    if not me or me.tenant_id != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Mass exam not found")

    result = await db.execute(
        select(Exam)
        .options(selectinload(Exam.employee).selectinload(Employee.user))
        .where(Exam.mass_exam_id == mass_exam_id, Exam.tenant_id == tenant_id)
        .order_by(Exam.created_at.desc())
    )
    exams = result.scalars().all()
    pass_mark = await _load_pass_mark(db, mass_exam_id)
    return [
        {
            "id": e.id,
            "employee_id": e.employee_id,
            "employee_name": _employee_display_name(e),
            # HRP-333: feed EmployeeSummaryLine in the Results table.
            # ``employee_status`` — ``status`` is the exam's own state.
            "position_title": e.employee.position_title if e.employee else None,
            "employee_status": e.employee.status if e.employee else None,
            "can_view_profile": (
                visible_employee_ids is None
                or e.employee_id in visible_employee_ids
            ),
            "status": e.status,
            "score": e.score,
            "max_score": e.max_score,
            "passed": _exam_passed(e.score, e.max_score, pass_mark),
            "started_at": e.started_at,
            "finished_at": e.finished_at,
        }
        for e in exams
    ]


def _employee_display_name(exam: Exam) -> str | None:
    emp = getattr(exam, "employee", None)
    if emp is None:
        return None
    user = getattr(emp, "user", None)
    if user is None:
        return None
    first = (user.first_name or "").strip()
    last = (user.last_name or "").strip()
    full = f"{first} {last}".strip()
    return full or (user.email or None)


# --- Individual Exam ---


async def list_my_exams(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    employee_id: uuid.UUID | None = None,
    visible_employee_ids: set[uuid.UUID] | None = None,
) -> list[dict]:
    query = (
        select(Exam)
        .options(selectinload(Exam.mass_exam))
        .where(Exam.tenant_id == tenant_id)
        .order_by(Exam.created_at.desc())
    )
    if employee_id is not None:
        query = query.where(Exam.employee_id == employee_id)
    elif visible_employee_ids is not None:
        if not visible_employee_ids:
            return []
        query = query.where(Exam.employee_id.in_(visible_employee_ids))
    result = await db.execute(query)
    rows = result.scalars().all()
    out: list[dict] = []
    # HRP-225: hide drafts from any scoped read — admin/HR without scope see
    # everything (no filter applied), but employees and managers calling
    # /exams pass `visible_employee_ids` (scoped to their own / subordinate
    # IDs) and must not see drafts that haven't been sent yet.
    scoped_read = employee_id is not None or visible_employee_ids is not None
    for e in rows:
        mass_exam = e.mass_exam
        if (
            scoped_read
            and mass_exam is not None
            and mass_exam.status == "draft"
        ):
            continue
        out.append(
            {
                "id": e.id,
                "mass_exam_id": e.mass_exam_id,
                "employee_id": e.employee_id,
                "status": e.status,
                "score": e.score,
                "max_score": e.max_score,
                "started_at": e.started_at,
                "finished_at": e.finished_at,
                "tenant_id": e.tenant_id,
                "created_at": e.created_at,
                "mass_exam_title": mass_exam.title if mass_exam else None,
            }
        )
    return out


def _assert_exam_owner(
    exam: Exam, acting_employee_id: uuid.UUID | None
) -> None:
    """HRP-328: taking endpoints are owner-only.

    ``acting_employee_id`` is None when called from service-level code that
    already established authorization (tests, seeds); the router always
    resolves and passes the caller's employee id.
    """
    if acting_employee_id is not None and exam.employee_id != acting_employee_id:
        raise HTTPException(
            status.HTTP_403_FORBIDDEN, "This exam belongs to another employee"
        )


async def get_exam_questions(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    exam_id: uuid.UUID,
    *,
    acting_employee_id: uuid.UUID | None = None,
) -> dict:
    """HRP-328: the take-sheet payload for the exam owner.

    Questions come WITHOUT ``is_correct`` (the employee must not see the
    answer key while taking the test), together with the answers saved so
    far so a reopened sheet resumes where the employee left off.
    """
    exam = await db.get(Exam, exam_id)
    if not exam or exam.tenant_id != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Exam not found")
    _assert_exam_owner(exam, acting_employee_id)

    mass_exam = await db.get(MassExam, exam.mass_exam_id)
    if mass_exam is None or mass_exam.status == "draft":
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Exam not found")
    if mass_exam.status == "cancelled" or exam.status == "cancelled":
        raise HTTPException(status.HTTP_409_CONFLICT, "Exam was cancelled")

    q_result = await db.execute(
        select(ExamQuestion)
        .options(selectinload(ExamQuestion.options))
        .where(
            ExamQuestion.mass_exam_id == exam.mass_exam_id,
            ExamQuestion.is_active.is_(True),
        )
        .order_by(ExamQuestion.sort_index, ExamQuestion.created_at)
    )
    questions = []
    for q in q_result.scalars().all():
        img_url = await _resolve_image_url(db, q.image_file_id)
        questions.append(
            {
                "id": q.id,
                "title": q.title,
                "description": q.description,
                "question_type": q.question_type,
                "sort_index": q.sort_index,
                "weight": q.weight,
                "image_url": img_url,
                "options": [
                    {"id": o.id, "title": o.title, "sort_index": o.sort_index}
                    for o in q.options
                ],
            }
        )

    a_result = await db.execute(
        select(ExamAnswer).where(ExamAnswer.exam_id == exam.id)
    )
    answers = [
        {
            "question_id": a.question_id,
            "selected_option_ids": a.selected_option_ids or [],
            "text_answer": a.text_answer,
        }
        for a in a_result.scalars().all()
    ]

    return {
        "exam_id": exam.id,
        "status": exam.status,
        "score": exam.score,
        "max_score": exam.max_score,
        "mass_exam_title": mass_exam.title,
        "questions": questions,
        "answers": answers,
    }


async def get_exam_review(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    exam_id: uuid.UUID,
    *,
    acting_employee_id: uuid.UUID | None = None,
    visible_employee_ids: set[uuid.UUID] | None = None,
) -> dict:
    """HRP-328: submitted-results view — the answer key is revealed.

    Available to the exam owner and to anyone whose access scope covers the
    employee (manager of the subtree, admin/HR). Only once the exam is done.
    """
    exam = await db.get(Exam, exam_id)
    if not exam or exam.tenant_id != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Exam not found")
    is_owner = (
        acting_employee_id is not None and exam.employee_id == acting_employee_id
    )
    in_scope = (
        visible_employee_ids is None or exam.employee_id in visible_employee_ids
    )
    if not is_owner and not in_scope:
        raise HTTPException(
            status.HTTP_403_FORBIDDEN, "This exam belongs to another employee"
        )
    # HRP-236: cancelling the whole procedure voids the survey status but
    # keeps submitted results reviewable — gate on the submission itself.
    if exam.status != "done" and exam.finished_at is None:
        if exam.status == "cancelled":
            # "available after completion" would be a lie — a voided
            # survey with no submission never gets results.
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                "The exam was cancelled before results were submitted",
            )
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            "Results are available after the exam is completed",
        )

    mass_exam = await db.get(MassExam, exam.mass_exam_id)
    q_result = await db.execute(
        select(ExamQuestion)
        .options(selectinload(ExamQuestion.options))
        .where(
            ExamQuestion.mass_exam_id == exam.mass_exam_id,
            ExamQuestion.is_active.is_(True),
        )
        .order_by(ExamQuestion.sort_index, ExamQuestion.created_at)
    )
    a_result = await db.execute(
        select(ExamAnswer).where(ExamAnswer.exam_id == exam.id)
    )
    answers_by_q = {a.question_id: a for a in a_result.scalars().all()}

    questions = []
    for q in q_result.scalars().all():
        img_url = await _resolve_image_url(db, q.image_file_id)
        ans = answers_by_q.get(q.id)
        questions.append(
            {
                "id": q.id,
                "title": q.title,
                "description": q.description,
                "question_type": q.question_type,
                "sort_index": q.sort_index,
                "weight": q.weight,
                "image_url": img_url,
                "options": [
                    {
                        "id": o.id,
                        "title": o.title,
                        "sort_index": o.sort_index,
                        "is_correct": o.is_correct,
                    }
                    for o in q.options
                ],
                "answer": (
                    {
                        "selected_option_ids": ans.selected_option_ids or [],
                        "text_answer": ans.text_answer,
                        "is_correct": ans.is_correct,
                        "score": ans.score,
                    }
                    if ans is not None
                    else None
                ),
            }
        )

    return {
        "exam_id": exam.id,
        "status": exam.status,
        "score": exam.score,
        "max_score": exam.max_score,
        "finished_at": exam.finished_at,
        "mass_exam_title": mass_exam.title if mass_exam else None,
        "questions": questions,
    }


async def submit_answer(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    exam_id: uuid.UUID,
    data,
    *,
    acting_employee_id: uuid.UUID | None = None,
) -> dict:

    exam = await db.get(Exam, exam_id)
    if not exam or exam.tenant_id != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Exam not found")
    _assert_exam_owner(exam, acting_employee_id)

    # HRP-328: no answers on finished exams, none before the send, and none
    # after the manager closed or cancelled the whole exam.
    if exam.status == "done":
        raise HTTPException(
            status.HTTP_409_CONFLICT, "Exam is already completed"
        )
    if exam.status == "cancelled":
        raise HTTPException(status.HTTP_409_CONFLICT, "Exam was cancelled")
    mass_exam = await db.get(MassExam, exam.mass_exam_id)
    if mass_exam is not None and mass_exam.status in {
        "draft",
        "done",
        "cancelled",
    }:
        raise HTTPException(
            status.HTTP_409_CONFLICT, "Exam is not open for answers"
        )

    if exam.status == "assigned":
        exam.status = "in_progress"
        exam.started_at = datetime.now(timezone.utc)

        # Auto-flip the mass exam to in_progress on the first participant's
        # first answer (the manager can't do this manually under the new
        # status model — HRP-236).
        if mass_exam is not None and mass_exam.status == "sent":
            mass_exam.status = "in_progress"
            if mass_exam.started_at is None:
                mass_exam.started_at = datetime.now(timezone.utc)

    # Load question to auto-score
    q_result = await db.execute(
        select(ExamQuestion)
        .options(selectinload(ExamQuestion.options))
        .where(ExamQuestion.id == data.question_id)
    )
    question = q_result.scalar_one_or_none()
    # HRP-328: the question must belong to THIS exam's test — otherwise an
    # answer could be injected from a foreign mass exam.
    if not question or question.mass_exam_id != exam.mass_exam_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Question not found")

    is_correct = None
    score = 0

    if (
        question.question_type in ("single_choice", "multiple_choice")
        and data.selected_option_ids
    ):
        correct_ids = {str(o.id) for o in question.options if o.is_correct}
        selected_ids = {str(sid) for sid in data.selected_option_ids}
        is_correct = correct_ids == selected_ids
        score = question.weight if is_correct else 0

    # HRP-328: upsert per (exam, question) — re-answering replaces the
    # previous row instead of stacking duplicates that would double-count
    # in finish_exam's sum. Backed by uq_exam_answer_exam_question so two
    # concurrent autosaves cannot both insert; the loser of the race
    # updates the winner's row via ON CONFLICT.
    selected_payload = (
        [str(sid) for sid in data.selected_option_ids]
        if data.selected_option_ids
        else None
    )
    values = {
        "exam_id": exam_id,
        "question_id": data.question_id,
        "selected_option_ids": selected_payload,
        "text_answer": data.text_answer,
        "is_correct": is_correct,
        "score": score,
    }
    stmt = pg_insert(ExamAnswer).values(**values)
    await db.execute(
        stmt.on_conflict_do_update(
            constraint="uq_exam_answer_exam_question",
            set_={
                "selected_option_ids": stmt.excluded.selected_option_ids,
                "text_answer": stmt.excluded.text_answer,
                "is_correct": stmt.excluded.is_correct,
                "score": stmt.excluded.score,
                "updated_at": func.now(),
            },
        )
    )
    await db.commit()

    return {"question_id": data.question_id, "is_correct": is_correct, "score": score}


# ---------------------------------------------------------------------------
# GF9: Exam Excel Import
# ---------------------------------------------------------------------------


async def import_questions_from_excel(
    db: AsyncSession, tenant_id: uuid.UUID, mass_exam_id: uuid.UUID, file_data: bytes
) -> dict:
    """Import exam questions from Excel.

    Expected columns: question_text, question_type, option_1, option_2, ..., option_6,
    correct_options (comma-separated indices 1-based), weight.
    """
    me = await db.get(MassExam, mass_exam_id)
    if not me or me.tenant_id != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Mass exam not found")

    wb = load_workbook(BytesIO(file_data), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    imported = 0
    errors: list[dict] = []
    seen_titles: set[str] = set()

    # Load existing question titles for duplicate detection
    existing = await db.execute(
        select(ExamQuestion.title).where(ExamQuestion.mass_exam_id == mass_exam_id)
    )
    for (t,) in existing.all():
        seen_titles.add(t.strip().lower())

    # Get current max sort_index
    max_idx_result = await db.execute(
        select(func.max(ExamQuestion.sort_index)).where(
            ExamQuestion.mass_exam_id == mass_exam_id
        )
    )
    sort_index = (max_idx_result.scalar() or 0) + 1

    for row_num, row in enumerate(rows, start=2):
        try:
            if not row or not row[0]:
                continue

            question_text = str(row[0]).strip()
            if not question_text:
                errors.append({"row": row_num, "error": "Empty question text"})
                continue

            # Duplicate check
            if question_text.lower() in seen_titles:
                errors.append(
                    {
                        "row": row_num,
                        "error": f"Duplicate question: {question_text[:50]}",
                    }
                )
                continue

            question_type = str(row[1]).strip().lower() if row[1] else "single_choice"
            if question_type not in ("single_choice", "multiple_choice", "essay"):
                errors.append(
                    {"row": row_num, "error": f"Invalid question type: {question_type}"}
                )
                continue

            # Parse options (columns 2-7, up to 6 options)
            options_raw = []
            for i in range(2, min(8, len(row))):
                if row[i] and str(row[i]).strip():
                    options_raw.append(str(row[i]).strip())

            # correct_options column (1-based indices, comma-separated)
            correct_col_idx = 8 if len(row) > 8 else None
            correct_indices: set[int] = set()
            if correct_col_idx and row[correct_col_idx]:
                try:
                    correct_indices = {
                        int(x.strip()) for x in str(row[correct_col_idx]).split(",")
                    }
                except ValueError:
                    errors.append(
                        {"row": row_num, "error": "Invalid correct_options format"}
                    )
                    continue

            # Weight column
            weight = 1
            weight_col_idx = 9 if len(row) > 9 else None
            if weight_col_idx and row[weight_col_idx]:
                try:
                    weight = int(float(str(row[weight_col_idx])))
                    if weight < 1:
                        weight = 1
                except ValueError:
                    weight = 1

            # Validate: choice questions need options
            if (
                question_type in ("single_choice", "multiple_choice")
                and len(options_raw) < 2
            ):
                errors.append(
                    {
                        "row": row_num,
                        "error": "Choice questions require at least 2 options",
                    }
                )
                continue

            # Create question
            q = ExamQuestion(
                mass_exam_id=mass_exam_id,
                title=question_text,
                question_type=question_type,
                sort_index=sort_index,
                weight=weight,
            )
            db.add(q)
            await db.flush()

            # Create options
            for opt_idx, opt_text in enumerate(options_raw):
                db.add(
                    QuestionOption(
                        question_id=q.id,
                        title=opt_text,
                        sort_index=opt_idx,
                        is_correct=(opt_idx + 1) in correct_indices,
                    )
                )

            seen_titles.add(question_text.lower())
            sort_index += 1
            imported += 1

        except Exception as e:  # noqa: BLE001 - per-row import isolation, recorded
            errors.append({"row": row_num, "error": str(e)})

    await db.commit()
    return {
        "mass_exam_id": mass_exam_id,
        "imported": imported,
        "errors": errors,
        "total_rows": len(rows),
    }


def generate_import_template() -> bytes:
    """Generate Excel template for question import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    headers = [
        "question_text",
        "question_type",
        "option_1",
        "option_2",
        "option_3",
        "option_4",
        "option_5",
        "option_6",
        "correct_options",
        "weight",
    ]
    ws.append(headers)
    # Example row
    ws.append(
        [
            "What is 2+2?",
            "single_choice",
            "3",
            "4",
            "5",
            "",
            "",
            "",
            "2",
            "1",
        ]
    )
    ws.append(
        [
            "Select prime numbers",
            "multiple_choice",
            "2",
            "3",
            "4",
            "5",
            "6",
            "",
            "1,2,4",
            "2",
        ]
    )
    ws.append(
        [
            "Explain the concept of polymorphism",
            "essay",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "3",
        ]
    )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def finish_exam(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    exam_id: uuid.UUID,
    *,
    acting_employee_id: uuid.UUID | None = None,
) -> dict:
    exam = await db.get(Exam, exam_id)
    if not exam or exam.tenant_id != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Exam not found")
    _assert_exam_owner(exam, acting_employee_id)
    if exam.status == "done":
        raise HTTPException(
            status.HTTP_409_CONFLICT, "Exam is already completed"
        )
    if exam.status == "cancelled":
        raise HTTPException(status.HTTP_409_CONFLICT, "Exam was cancelled")

    # HRP-328: results can only be submitted once every active question has
    # an answer (an essay counts when its text is non-empty).
    q_result = await db.execute(
        select(ExamQuestion).where(
            ExamQuestion.mass_exam_id == exam.mass_exam_id,
            ExamQuestion.is_active.is_(True),
        )
    )
    active_questions = q_result.scalars().all()
    a_result = await db.execute(select(ExamAnswer).where(ExamAnswer.exam_id == exam_id))
    answers = a_result.scalars().all()
    answered_ids = {
        a.question_id
        for a in answers
        if a.selected_option_ids or (a.text_answer or "").strip()
    }
    unanswered = [q for q in active_questions if q.id not in answered_ids]
    if unanswered:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Answer all questions before submitting ({len(unanswered)} left)",
        )

    # Calculate total score
    total_score = sum(a.score for a in answers)

    exam.score = total_score
    exam.status = "done"
    exam.finished_at = datetime.now(timezone.utc)

    # HRP-328: when the last participant finishes, the mass exam flips to
    # done automatically (the manager no longer has to close it by hand).
    # FOR UPDATE serializes two participants finishing at the same moment —
    # otherwise each could see the other as still unfinished and neither
    # would flip the parent.
    mass_exam = (
        await db.execute(
            select(MassExam)
            .where(MassExam.id == exam.mass_exam_id)
            .with_for_update()
            .execution_options(populate_existing=True)
        )
    ).scalar_one_or_none()
    # HRP-236 (review): the lock also serializes against a concurrent
    # cancel/complete of the whole exam. If the procedure went terminal
    # while this submission was in flight, the cascade already voided this
    # survey — committing "done" here would overwrite that cancel.
    if mass_exam is not None and mass_exam.status in {"done", "cancelled"}:
        raise HTTPException(status.HTTP_409_CONFLICT, "Exam was cancelled")
    if mass_exam is not None and mass_exam.status in {"sent", "in_progress"}:
        remaining = await db.scalar(
            select(func.count(Exam.id)).where(
                Exam.mass_exam_id == exam.mass_exam_id,
                Exam.id != exam.id,
                Exam.status != "done",
            )
        )
        if not remaining:
            mass_exam.status = "done"
            mass_exam.finished_at = datetime.now(timezone.utc)

    await db.commit()

    pass_mark = await _load_pass_mark(db, exam.mass_exam_id)
    passed = _exam_passed(total_score, exam.max_score, pass_mark) or False

    return {
        "exam_id": exam.id,
        "score": total_score,
        "max_score": exam.max_score,
        "passed": passed,
    }
