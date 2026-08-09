"""Report service — templates, consolidated XLSX reports, candidate compare.

Extracted from ``service.py`` during the M-5 split (R4d). Audit/billing
hooks target this module; ``service.py`` re-exports via module
``__getattr__`` for legacy callers.
"""

from __future__ import annotations

import uuid

from fastapi import status
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.errors import AppError
from app.modules.auth.models import User
from app.modules.recruitment.common import (
    _FULL_ROLES,
    _get_vacancy,
    candidate_display_name,
    normalize_competence_id,
)
from app.modules.recruitment.models import (
    AIAssessment,
    Candidate,
    CandidateVacancy,
    ConsolidatedReport,
    HumanAssessment,
    Interview,
    VacancyProfile,
)
from app.modules.recruitment.schemas import (
    REPORT_SECTION_CODES,
    ReportGenerateRequest,
)
from app.modules.storage.models import File

# ---------------------------------------------------------------------------
# Reports (R4a) — see RECRUITING_MODULE.md §11.2 SCR-80..83.
# HRP-521 removed the report-template surface: the sheet set now comes
# straight from the Generate-report dialog.
# ---------------------------------------------------------------------------


_DEFAULT_REPORT_SECTIONS: list[str] = list(REPORT_SECTION_CODES)


def _validate_sections(sections: list[str] | None) -> list[str]:
    if not sections:
        return list(_DEFAULT_REPORT_SECTIONS)
    seen: set[str] = set()
    cleaned: list[str] = []
    for s in sections:
        code = str(s).strip().lower()
        if code in REPORT_SECTION_CODES and code not in seen:
            cleaned.append(code)
            seen.add(code)
    return cleaned or list(_DEFAULT_REPORT_SECTIONS)


def _report_export_to_read(
    export: ConsolidatedReport,
    *,
    requested_by_name: str | None = None,
    download_url: str | None = None,
) -> dict:
    return {
        "id": export.id,
        "vacancy_id": export.vacancy_id,
        "sections": list(export.sections or []),
        "status": export.status,
        "error": export.error,
        "file_id": export.file_id,
        "download_url": download_url,
        "requested_by": export.generated_by,
        "requested_by_name": requested_by_name,
        "created_at": export.created_at,
        "completed_at": export.completed_at,
    }


async def _name_lookup_for_exports(
    db: AsyncSession, exports: list[ConsolidatedReport]
) -> dict[uuid.UUID, str]:
    user_ids = {e.generated_by for e in exports if e.generated_by}
    user_names: dict[uuid.UUID, str] = {}
    if user_ids:
        user_rows = (
            (await db.execute(select(User).where(User.id.in_(user_ids))))
            .scalars()
            .all()
        )
        user_names = {
            u.id: f"{u.first_name or ''} {u.last_name or ''}".strip() or u.email
            for u in user_rows
        }
    return user_names


async def enqueue_report(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    user_id: uuid.UUID,
    vacancy_id: uuid.UUID,
    data: ReportGenerateRequest,
) -> dict:
    """Persist a pending export row and dispatch the Celery task.

    Returns ``{export_id, task_id}`` so the frontend can poll until the
    XLSX file is ready (file_id populated, status='completed').
    """

    from app.modules.recruitment.tasks import generate_report_task

    await _get_vacancy(db, tenant_id, vacancy_id)

    sections: list[str] | None = (
        [str(s) for s in data.sections] if data.sections else None
    )
    sections_clean = _validate_sections(sections)

    # HRP-268 — stash the audience + candidate whitelist inside the
    # existing ``report_data`` JSONB so the Celery task picks them up
    # without a schema migration. Audience defaults to "recruiter" so
    # the Detail sheet shows raw process findings; "hiring_manager"
    # replaces them with the positive reframe.
    audience = (data.audience or "recruiter").lower()
    if audience not in {"recruiter", "hiring_manager"}:
        audience = "recruiter"
    report_data: dict = {"audience": audience}
    # Omitting the field means "every active candidate" (the wizard's
    # "All active" scope sends no whitelist). Sending an empty list is a
    # different statement: the caller picked a scope — Finalists, Custom —
    # that resolved to nobody. Truthiness collapsed the two, so an empty
    # selection silently produced a report on the whole vacancy, which is
    # the opposite of what was asked for. The wizard disables its button
    # in that state; the server says so too.
    if data.candidate_vacancy_ids is not None:
        if not data.candidate_vacancy_ids:
            raise AppError(
                "report_candidate_selection_empty", status.HTTP_400_BAD_REQUEST
            )
        report_data["candidate_vacancy_ids"] = [
            str(cv_id) for cv_id in data.candidate_vacancy_ids
        ]

    export = ConsolidatedReport(
        tenant_id=tenant_id,
        vacancy_id=vacancy_id,
        sections=sections_clean,
        status="pending",
        generated_by=user_id,
        report_data=report_data,
    )
    db.add(export)
    await db.commit()
    await db.refresh(export)

    task = generate_report_task.delay(str(export.id), str(tenant_id))
    return {"export_id": export.id, "task_id": task.id}


async def list_reports(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    *,
    vacancy_id: uuid.UUID | None = None,
    status_filter: str | None = None,
    skip: int = 0,
    limit: int = 50,
) -> tuple[list[dict], int]:
    query = select(ConsolidatedReport).where(ConsolidatedReport.tenant_id == tenant_id)
    count_query = select(func.count(ConsolidatedReport.id)).where(
        ConsolidatedReport.tenant_id == tenant_id
    )
    if vacancy_id is not None:
        query = query.where(ConsolidatedReport.vacancy_id == vacancy_id)
        count_query = count_query.where(ConsolidatedReport.vacancy_id == vacancy_id)
    if status_filter:
        query = query.where(ConsolidatedReport.status == status_filter)
        count_query = count_query.where(ConsolidatedReport.status == status_filter)

    total = (await db.execute(count_query)).scalar() or 0
    rows = (
        (
            await db.execute(
                query.order_by(ConsolidatedReport.created_at.desc())
                .offset(skip)
                .limit(limit)
            )
        )
        .scalars()
        .all()
    )
    user_names = await _name_lookup_for_exports(db, list(rows))
    items = [
        _report_export_to_read(
            r,
            requested_by_name=user_names.get(r.generated_by)
            if r.generated_by
            else None,
        )
        for r in rows
    ]
    return items, total


async def get_report(
    db: AsyncSession, tenant_id: uuid.UUID, export_id: uuid.UUID
) -> dict:
    export = (
        await db.execute(
            select(ConsolidatedReport).where(
                ConsolidatedReport.id == export_id,
                ConsolidatedReport.tenant_id == tenant_id,
            )
        )
    ).scalar_one_or_none()
    if not export:
        raise AppError("report_not_found", status.HTTP_404_NOT_FOUND)

    requested_by_name: str | None = None
    if export.generated_by:
        usr = (
            await db.execute(select(User).where(User.id == export.generated_by))
        ).scalar_one_or_none()
        if usr:
            requested_by_name = (
                f"{usr.first_name or ''} {usr.last_name or ''}".strip() or usr.email
            )

    download_url: str | None = None
    if export.status == "completed" and export.file_id:
        from app.core.s3 import get_presigned_url

        file_row = (
            await db.execute(select(File).where(File.id == export.file_id))
        ).scalar_one_or_none()
        if file_row:
            # HRP-268 — bumped to 24h so a recruiter who emails the
            # link to a hiring manager can still open it the next day.
            download_url = get_presigned_url(file_row.path, expires_in=86400)

    return _report_export_to_read(
        export,
        requested_by_name=requested_by_name,
        download_url=download_url,
    )


async def delete_report(
    db: AsyncSession, tenant_id: uuid.UUID, export_id: uuid.UUID
) -> None:
    export = (
        await db.execute(
            select(ConsolidatedReport).where(
                ConsolidatedReport.id == export_id,
                ConsolidatedReport.tenant_id == tenant_id,
            )
        )
    ).scalar_one_or_none()
    if not export:
        raise AppError("report_not_found", status.HTTP_404_NOT_FOUND)

    file_path: str | None = None
    file_row: File | None = None
    if export.file_id:
        file_row = (
            await db.execute(select(File).where(File.id == export.file_id))
        ).scalar_one_or_none()
        if file_row:
            file_path = file_row.path

    await db.delete(export)
    if file_row is not None:
        await db.delete(file_row)
    await db.commit()

    if file_path:
        # Best-effort S3 delete — a missing object must not fail the API.
        try:
            from app.core.s3 import delete_file

            delete_file(file_path)
        except Exception:  # noqa: BLE001 - best-effort S3 delete
            pass


# ---------------------------------------------------------------------------
# XLSX preview (R4d, FR-23 / SCR-83)
# ---------------------------------------------------------------------------


# Per-sheet caps so a malicious or accidentally enormous XLSX cannot pin
# backend memory or blow up the JSON payload the frontend has to render.
# 200 rows × 50 cols ≈ 10k cells which fits comfortably in a single
# in-browser table. Anything bigger triggers ``truncated=True``.
_PREVIEW_MAX_ROWS = 200
_PREVIEW_MAX_COLS = 50

# Redis TTL for the preview cache. Five minutes matches the user-side
# cadence: a recruiter usually opens the same report a couple times in a
# session, but the underlying file does not change so a longer TTL would
# only hide regenerations.
_PREVIEW_CACHE_TTL_SECONDS = 300


def _preview_cache_key(tenant_id: uuid.UUID, export_id: uuid.UUID) -> str:
    return f"recruitment:report_preview:{tenant_id}:{export_id}"


async def get_report_preview(
    db: AsyncSession, tenant_id: uuid.UUID, export_id: uuid.UUID
) -> dict:
    """Return a JSON-friendly snapshot of every sheet in the report XLSX.

    The XLSX itself stays in S3; we stream it into memory and convert
    each sheet into a list of rows so the frontend can render it inline
    without a download. Results are cached in Redis for five minutes —
    the recruiter typically opens the same report several times in a
    session and the file never changes after ``status='completed'``.
    """

    import json
    import logging

    logger = logging.getLogger(__name__)

    export = (
        await db.execute(
            select(ConsolidatedReport).where(
                ConsolidatedReport.id == export_id,
                ConsolidatedReport.tenant_id == tenant_id,
            )
        )
    ).scalar_one_or_none()
    if not export:
        raise AppError("report_not_found", status.HTTP_404_NOT_FOUND)
    if export.status != "completed" or export.file_id is None:
        raise AppError("report_not_ready", status.HTTP_409_CONFLICT)

    cache_key = _preview_cache_key(tenant_id, export_id)
    redis_client = None
    try:
        import redis.asyncio as aioredis

        from app.config import settings

        redis_client = aioredis.from_url(settings.redis_url, decode_responses=True)
        cached = await redis_client.get(cache_key)
        if cached:
            try:
                return json.loads(cached)
            except json.JSONDecodeError:
                pass
    except Exception:  # noqa: BLE001
        # Redis is best-effort here. If it's down we just re-parse the
        # XLSX every time — slower, never wrong.
        redis_client = None

    file_row = (
        await db.execute(select(File).where(File.id == export.file_id))
    ).scalar_one_or_none()
    if not file_row:
        raise AppError("report_file_missing", status.HTTP_404_NOT_FOUND)

    from app.config import settings
    from app.core.s3 import get_s3_client

    client = get_s3_client()
    if client is None:
        raise AppError(
            "object_storage_not_configured", status.HTTP_503_SERVICE_UNAVAILABLE
        )

    import io

    try:
        obj = client.get_object(Bucket=settings.s3_bucket, Key=file_row.path)
        body = obj["Body"].read()
    except Exception as exc:  # noqa: BLE001
        logger.exception("report preview: S3 fetch failed for %s", file_row.path)
        raise AppError(
            "report_storage_read_failed", status.HTTP_503_SERVICE_UNAVAILABLE
        ) from exc

    from openpyxl import load_workbook

    sheets: list[dict] = []
    truncated = False
    try:
        wb = load_workbook(filename=io.BytesIO(body), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.exception("report preview: openpyxl failed to parse %s", file_row.path)
        raise AppError(
            "report_file_not_readable_xlsx", status.HTTP_422_UNPROCESSABLE_ENTITY
        ) from exc

    try:
        for ws in wb.worksheets:
            rows_out: list[list] = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx >= _PREVIEW_MAX_ROWS:
                    truncated = True
                    break
                trimmed = list(row[:_PREVIEW_MAX_COLS])
                if len(row) > _PREVIEW_MAX_COLS:
                    truncated = True
                # JSON-safe coercion: dates / Decimals → strings.
                cleaned: list = []
                for v in trimmed:
                    if v is None or isinstance(v, (str, int, float, bool)):
                        cleaned.append(v)
                    else:
                        cleaned.append(str(v))
                rows_out.append(cleaned)
            sheets.append({"name": ws.title, "cells": rows_out})
    finally:
        wb.close()

    payload = {
        "export_id": str(export_id),
        "sheets": sheets,
        "truncated": truncated,
    }

    if redis_client is not None:
        import contextlib

        with contextlib.suppress(Exception):
            await redis_client.set(
                cache_key, json.dumps(payload), ex=_PREVIEW_CACHE_TTL_SECONDS
            )

    return payload


# ---------------------------------------------------------------------------
# Candidate comparison (R4a, SCR-55/56)
# ---------------------------------------------------------------------------


_MAX_COMPARISON_CANDIDATES = 5


async def compare_candidates(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    vacancy_id: uuid.UUID,
    candidate_ids: list[uuid.UUID],
    *,
    role: str | None = None,
) -> dict:
    """Build the side-by-side competence matrix used by SCR-55/56.

    For each competence in the vacancy profile we collect human +
    AI scores per candidate and compute ``divergence = max - min`` over
    the cells that have a numeric score. ``role`` controls whether
    hiring-manager-shaped output is returned (no AI reasoning, no
    red-flag-style detail).
    """

    if not candidate_ids:
        raise AppError("report_candidate_ids_required", status.HTTP_400_BAD_REQUEST)
    unique_ids: list[uuid.UUID] = []
    seen: set[uuid.UUID] = set()
    for cid in candidate_ids:
        if cid not in seen:
            unique_ids.append(cid)
            seen.add(cid)
    if len(unique_ids) > _MAX_COMPARISON_CANDIDATES:
        raise AppError(
            "report_too_many_comparison_candidates",
            status.HTTP_400_BAD_REQUEST,
            max_candidates=_MAX_COMPARISON_CANDIDATES,
        )

    await _get_vacancy(db, tenant_id, vacancy_id)

    profile = (
        await db.execute(
            select(VacancyProfile).where(
                VacancyProfile.vacancy_id == vacancy_id,
                VacancyProfile.tenant_id == tenant_id,
            )
        )
    ).scalar_one_or_none()
    profile_competences = (
        (profile.profile_data or {}).get("competences", []) if profile else []
    )

    cvs = (
        (
            await db.execute(
                select(CandidateVacancy)
                .options(
                    selectinload(CandidateVacancy.candidate).selectinload(
                        Candidate.person
                    )
                )
                .where(
                    CandidateVacancy.vacancy_id == vacancy_id,
                    CandidateVacancy.tenant_id == tenant_id,
                    CandidateVacancy.candidate_id.in_(unique_ids),
                )
            )
        )
        .scalars()
        .unique()
        .all()
    )

    if not cvs:
        return {
            "vacancy_id": vacancy_id,
            "candidates": [],
            "competences": [],
        }

    cv_ids = [cv.id for cv in cvs]

    human_rows = (
        (
            await db.execute(
                select(HumanAssessment).where(
                    HumanAssessment.candidate_vacancy_id.in_(cv_ids),
                    HumanAssessment.tenant_id == tenant_id,
                )
            )
        )
        .scalars()
        .all()
    )

    interviews = (
        (
            await db.execute(
                select(Interview).where(
                    Interview.candidate_vacancy_id.in_(cv_ids),
                    Interview.tenant_id == tenant_id,
                )
            )
        )
        .scalars()
        .all()
    )
    interview_cv_map = {iv.id: iv.candidate_vacancy_id for iv in interviews}
    ai_rows: list[AIAssessment] = []
    if interviews:
        ai_rows = list(
            (
                await db.execute(
                    select(AIAssessment).where(
                        AIAssessment.interview_id.in_([iv.id for iv in interviews]),
                        AIAssessment.tenant_id == tenant_id,
                    )
                )
            )
            .scalars()
            .all()
        )

    # candidate_id → cv (one row per candidate on this vacancy)
    cv_by_candidate: dict[uuid.UUID, CandidateVacancy] = {
        cv.candidate_id: cv for cv in cvs
    }

    # Average scores per candidate per competence (averaging across
    # evaluators on the human side).
    human_scores: dict[uuid.UUID, dict[str, list[float]]] = {}
    for hs in human_rows:
        if hs.score is None:
            continue
        target = next((cv for cv in cvs if cv.id == hs.candidate_vacancy_id), None)
        if target is None:
            continue
        cand_id = target.candidate_id
        comp_id = str(hs.competence_id)
        bucket = human_scores.setdefault(cand_id, {}).setdefault(comp_id, [])
        bucket.append(float(hs.score))

    ai_scores: dict[uuid.UUID, dict[str, list[float]]] = {}
    for ai in ai_rows:
        if ai.score is None:
            continue
        cv_id = interview_cv_map.get(ai.interview_id)
        if cv_id is None:
            continue
        target = next((cv for cv in cvs if cv.id == cv_id), None)
        if target is None:
            continue
        cand_id = target.candidate_id
        comp_id = str(ai.competence_id)
        bucket = ai_scores.setdefault(cand_id, {}).setdefault(comp_id, [])
        bucket.append(float(ai.score))

    role_norm = (role or "").lower()
    # Defence in depth: only the four "full" recruitment roles see AI
    # scores. Hiring managers + unknown roles fall back to "human only"
    # — same posture as `_role_filter_analysis`, which returns None for
    # unknown roles. The router still gates with `get_current_user` so
    # callers must be in the tenant; this layer hardens the read shape.
    hide_ai = role_norm not in _FULL_ROLES

    candidates_payload: list[dict] = []
    for cid in unique_ids:
        cv = cv_by_candidate.get(cid)
        if cv is None:
            continue
        # HRP-361: full_name-first — resume-sourced candidates have no
        # Person row and used to render as "Unknown" in reports.
        name = candidate_display_name(cv.candidate)
        candidates_payload.append(
            {
                "id": cid,
                "candidate_vacancy_id": cv.id,
                "name": name or "Unknown",
                "status": cv.status,
                "ranking_score": cv.ranking_score,
            }
        )

    # Build competence rows. Slugged ids in the profile go through
    # normalize_competence_id so they match what scoring writes.
    competence_payload: list[dict] = []
    for comp in profile_competences or []:
        if not isinstance(comp, dict):
            continue
        raw_id = comp.get("id") or comp.get("name") or ""
        comp_uuid = normalize_competence_id(raw_id)
        if comp_uuid is None:
            continue
        comp_id_str = str(comp_uuid)
        cells: list[dict] = []
        scored_values: list[float] = []
        for cand in candidates_payload:
            cand_id = cand["id"]
            human_vals = human_scores.get(cand_id, {}).get(comp_id_str)
            score: float | None = None
            source = "missing"
            if human_vals:
                score = sum(human_vals) / len(human_vals)
                source = "human"
            elif not hide_ai:
                ai_vals = ai_scores.get(cand_id, {}).get(comp_id_str)
                if ai_vals:
                    score = sum(ai_vals) / len(ai_vals)
                    source = "ai"
            if score is not None:
                scored_values.append(score)
            cells.append(
                {
                    "candidate_id": cand_id,
                    "score": score,
                    "source": source,
                }
            )
        divergence: float | None = None
        if len(scored_values) >= 2:
            divergence = round(max(scored_values) - min(scored_values), 2)
        competence_payload.append(
            {
                "competence_id": comp_id_str,
                "name": comp.get("name") or comp.get("id") or "",
                "group": comp.get("group"),
                "criticality": comp.get("criticality"),
                "cells": cells,
                "divergence": divergence,
            }
        )

    return {
        "vacancy_id": vacancy_id,
        "candidates": candidates_payload,
        "competences": competence_payload,
    }
