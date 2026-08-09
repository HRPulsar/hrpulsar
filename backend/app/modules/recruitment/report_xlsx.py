"""Consolidated XLSX renderer (HRP-268, FR-23 / FR-24).

The workbook carries up to four kinds of sheet, in this order — which of
them are emitted is chosen in the Generate-report dialog (HRP-521):

1. **Summary** — one row per candidate with profile snippet, manager %,
   AI %, AI data completeness and a computed recommendation.
2. **Matrix** — competences × candidates with side-by-side Manager / AI
   columns, group separators, amber divergence highlight and a SUM-based
   "Total" row.
3. **Detail · {candidate}** — one sheet per candidate carrying resume
   summary, per-competence scoring with citations, blind spots, red
   flags, process findings (reframed for the hiring-manager audience),
   and verdict.
4. **Incomplete data** — candidates whose AI / manager scoring is not
   ready yet, with a "what to do" hint. Rendered even when empty.

Every sheet reserves row 1 for the tenant logo (HRP-522), so titles
start on row 2.

The renderer is intentionally stateless: callers assemble the payload
in the Celery task so the module is exercisable in unit tests without
DB / S3 access.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Style palette — kept tight so a future visual refresh changes one place.
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
_HEADER_FILL = PatternFill(start_color="1A1F36", end_color="1A1F36", fill_type="solid")
_GROUP_FILL = PatternFill(start_color="F1F3F8", end_color="F1F3F8", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
_AMBER_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
_EMERALD_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
_ROSE_FILL = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")

# HRP-524 REDO — a Matrix cell where Manager and AI diverge carries the ⚠
# marker in its *number format*, not in its value. Writing "2.0 ⚠" made the
# cell a string, and the Total row is a live ``=SUM()`` that silently skips
# text: the candidate with the most divergences got the lowest total, while
# the Python-side threshold fill (computed from the real percentage) said
# the opposite. The number format keeps the cell numeric — SUM counts it —
# and still renders the marker to the reader.
_DIVERGENCE_FORMAT = '0.0" ⚠"'

_HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
_BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BOLD = Font(name="Calibri", size=11, bold=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)


def _safe_sheet_title(title: str, *, reserve: int = 0) -> str:
    """Excel sheet titles cap at 31 chars and forbid ``: \\ / ? * [ ]``.

    ``reserve`` keeps the last N characters free so :func:`render_report_xlsx`
    can suffix a dedup counter (`#2`, `#3`) on collision without ever
    exceeding the 31-char limit.
    """
    cleaned = "".join(ch for ch in title if ch not in ":\\/?*[]")
    limit = max(1, 31 - reserve)
    return cleaned[:limit] or "Sheet"


def _coerce(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _set_header_cell(ws: Worksheet, row: int, col: int, value: Any) -> None:
    cell = ws.cell(row=row, column=col, value=_coerce(value))
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = _HEADER_ALIGN
    cell.border = _THIN_BORDER


def _set_body_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value: Any,
    *,
    align: Alignment | None = None,
    fill: PatternFill | None = None,
    bold: bool = False,
    number_format: str | None = None,
) -> None:
    cell = ws.cell(row=row, column=col, value=_coerce(value))
    cell.alignment = align or _BODY_ALIGN
    if fill is not None:
        cell.fill = fill
    if bold:
        cell.font = _BOLD
    if number_format is not None:
        cell.number_format = number_format
    cell.border = _THIN_BORDER


# HRP-522 — the logo sits in the top-left corner of every sheet at the
# standard 120x40 px. Row 1 is reserved for it (40 px ~ 30 pt) and the
# title starts on row 2, so the image never covers text. The reservation
# is unconditional: sheet geometry must not depend on whether a tenant
# happens to have uploaded a logo.
_LOGO_WIDTH_PX = 120
_LOGO_HEIGHT_PX = 40
_LOGO_ROW_HEIGHT_PT = 32

_TITLE_ROW = 2
_SUBTITLE_ROW = 3


def _embed_logo(ws: Worksheet, branding: dict | None, anchor: str = "A1") -> None:
    """Place the tenant logo in the sheet's top-left corner.

    A tenant without a logo simply gets no image — never an exception,
    never a broken export.
    """
    ws.row_dimensions[1].height = _LOGO_ROW_HEIGHT_PT
    if not branding or not isinstance(branding.get("logo_png"), bytes):
        return
    try:
        from openpyxl.drawing.image import Image as XLImage

        buf = io.BytesIO(branding["logo_png"])
        img = XLImage(buf)
        img.width = _LOGO_WIDTH_PX
        img.height = _LOGO_HEIGHT_PX
        ws.add_image(img, anchor)
    except Exception:  # pragma: no cover  # noqa: BLE001 - never blocks export
        return


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def render_report_xlsx(
    *,
    vacancy: dict,
    branding: dict | None = None,
    summary_rows: list[dict] | None = None,
    matrix: dict | None = None,
    details: list[dict] | None = None,
    incomplete: list[dict] | None = None,
    audience: str | None = None,
    sections: list[str] | None = None,
) -> bytes:
    """Build the consolidated workbook and return raw XLSX bytes.

    ``vacancy`` carries ``title``, ``specialization_title``,
    ``division_name`` etc. for the headers across sheets. ``audience``
    is ``recruiter`` (default) or ``hiring_manager`` — the latter
    replaces the raw process-findings text on each Detail sheet with
    the ``positive_reframe`` field, falling back to "Recommendation for
    the next interview" when missing.

    ``sections`` (HRP-521) picks which sheets to emit — the codes the
    Generate-report dialog sends. ``None`` means all four. A workbook
    must contain at least one sheet, so an empty / unknown selection
    falls back to the full set rather than producing an invalid file.
    """

    known = (
        "summary_ranking",
        "competency_matrix",
        "detailed_analysis",
        "incomplete_data",
    )
    wanted = {str(code) for code in sections} & set(known) if sections else set(known)
    if not wanted:
        wanted = set(known)

    wb = Workbook()
    default_ws: Worksheet | None = wb.active
    used_titles: set[str] = set()

    def _take_sheet(title: str, *, reserve: int = 0) -> Worksheet:
        nonlocal default_ws
        # Dedup on collision (two candidates sharing the first ~22 chars
        # of name) by appending ` #N` *inside* the 31-char budget. The
        # reserve keeps room for the suffix so we never produce a
        # >31-char title that openpyxl warns about + Excel may reject.
        base = _safe_sheet_title(title, reserve=reserve)
        final = base
        counter = 2
        while final in used_titles:
            suffix = f" #{counter}"
            final = _safe_sheet_title(title, reserve=len(suffix)) + suffix
            counter += 1
        used_titles.add(final)
        if default_ws is not None:
            ws = default_ws
            ws.title = final
            default_ws = None
            return ws
        return wb.create_sheet(title=final)

    if "summary_ranking" in wanted:
        _render_summary(
            _take_sheet("Summary"), vacancy, summary_rows or [], branding=branding
        )
    if "competency_matrix" in wanted:
        _render_matrix(_take_sheet("Matrix"), matrix or {}, vacancy, branding=branding)
    if "detailed_analysis" in wanted:
        for detail in details or []:
            title = "Detail · " + (detail.get("candidate_name") or "Candidate")
            # Reserve 4 chars (` #99`) so the dedup counter cannot push the
            # final title past the 31-char limit even for the longest names.
            _render_detail(
                _take_sheet(title, reserve=4),
                detail,
                audience=audience or "recruiter",
                branding=branding,
            )
    if "incomplete_data" in wanted:
        _render_incomplete(
            _take_sheet("Incomplete data"), incomplete or [], branding=branding
        )

    # A selection that produced no sheet at all (e.g. Detail-only with no
    # candidates) would leave openpyxl's default sheet dangling — keep it
    # as an explicit, self-explaining placeholder instead of saving a
    # workbook Excel refuses to open.
    if default_ws is not None:
        ws = default_ws
        ws.title = "Report"
        ws.cell(row=1, column=1, value="No data matched the selected sheets.")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Sheet 1 — Summary ranking
# ---------------------------------------------------------------------------


def _render_summary(
    ws: Worksheet, vacancy: dict, rows: list[dict], *, branding: dict | None
) -> None:
    """Sheet 1 (HRP-523) — the candidate ranking table.

    Layout: logo band, title, position/division/date line, an
    auto-generated data-completeness disclaimer, then one row per
    candidate. Footnotes for partial AI coverage / manager-AI divergence
    are collected under a "Notes" block and referenced by an asterisk on
    the AI score cell.
    """

    _embed_logo(ws, branding)
    tenant_name = (branding or {}).get("tenant_name") or ""
    title_line = (
        f"Candidate ranking — {vacancy.get('title') or 'Vacancy'}"
        f"{' | ' + tenant_name if tenant_name else ''}"
    )
    ws.cell(row=_TITLE_ROW, column=1, value=title_line).font = Font(size=14, bold=True)

    subtitle_bits = [
        f"Position: {vacancy.get('specialization_title')}"
        if vacancy.get("specialization_title")
        else None,
        vacancy.get("division_name"),
        f"Report date: {_today()}",
    ]
    ws.cell(
        row=_SUBTITLE_ROW,
        column=1,
        value=" | ".join([b for b in subtitle_bits if b]),
    ).font = Font(size=10, color="6B7280")

    disclaimer = _completeness_disclaimer(rows)
    if disclaimer:
        cell = ws.cell(row=_SUBTITLE_ROW + 1, column=1, value=disclaimer)
        cell.font = Font(size=10, color="92400E")
        cell.alignment = _BODY_ALIGN

    headers = [
        "Candidate",
        "Profile",
        "Manager score",
        "AI score",
        "AI data",
        "Recommendation",
    ]
    header_row = 6
    for col, value in enumerate(headers, start=1):
        _set_header_cell(ws, header_row, col, value)
    ws.row_dimensions[header_row].height = 22

    for idx, row in enumerate(rows, start=header_row + 1):
        recommendation = row.get("recommendation") or "Additional check"
        rec_fill = _recommendation_fill(recommendation)
        _set_body_cell(ws, idx, 1, row.get("candidate_name"))
        _set_body_cell(ws, idx, 2, row.get("profile_summary"))
        _set_body_cell(ws, idx, 3, row.get("manager_score_text"))
        _set_body_cell(ws, idx, 4, row.get("ai_score_text"))
        _set_body_cell(ws, idx, 5, row.get("data_readiness"))
        _set_body_cell(ws, idx, 6, recommendation, fill=rec_fill, bold=True)

    notes = [r for r in rows if r.get("note")]
    if notes:
        notes_row = header_row + len(rows) + 2
        ws.cell(row=notes_row, column=1, value="NOTES").font = _BOLD
        for offset, row in enumerate(notes, start=notes_row + 1):
            _set_body_cell(
                ws,
                offset,
                1,
                f"* {row.get('candidate_name')}: {row.get('note')}",
            )

    _autosize(ws, headers, default=24)
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 28


def _today() -> str:
    from datetime import datetime, timezone

    return datetime.now(timezone.utc).date().isoformat()


def _completeness_disclaimer(rows: list[dict]) -> str:
    """One line summarising how trustworthy the AI column is.

    Counts candidates by ``data_readiness`` so the reader knows up front
    how much of the table rests on transcript-backed analysis.
    """

    if not rows:
        return ""
    total = len(rows)
    full = sum(1 for r in rows if r.get("data_readiness") == "Full data")
    no_transcript = sum(1 for r in rows if r.get("data_readiness") == "No transcript")
    no_data = sum(1 for r in rows if r.get("data_readiness") == "No data")
    if full == total:
        return (
            "⚠ All candidates have resume and transcript on file — "
            "AI scores are directly comparable."
        )
    parts = [f"{full} of {total} candidates have full data"]
    if no_transcript:
        parts.append(f"{no_transcript} have no interview transcript")
    if no_data:
        parts.append(f"{no_data} have no AI analysis at all")
    return (
        "⚠ "
        + "; ".join(parts)
        + ". AI scores are not comparable across rows with different data "
        "readiness."
    )


def _recommendation_fill(value: str) -> PatternFill | None:
    lowered = (value or "").lower()
    if "not" in lowered or "reject" in lowered:
        return _ROSE_FILL
    if "review" in lowered or "additional" in lowered or "check" in lowered:
        return _AMBER_FILL
    if "recommend" in lowered or "hire" in lowered:
        return _EMERALD_FILL
    return None


# ---------------------------------------------------------------------------
# Sheet 2 — Competence matrix
# ---------------------------------------------------------------------------


def _render_matrix(
    ws: Worksheet, matrix: dict, vacancy: dict, *, branding: dict | None
) -> None:
    _embed_logo(ws, branding)
    competences: list[dict] = matrix.get("competences") or []
    candidates: list[dict] = matrix.get("candidates") or []
    max_score: float = float(matrix.get("max_score") or 5.0)

    vacancy_title = matrix.get("vacancy_title") or vacancy.get("title") or "Vacancy"
    title = f"Competence matrix — {len(candidates)} candidates | {vacancy_title}"
    ws.cell(row=_TITLE_ROW, column=1, value=title).font = Font(size=14, bold=True)
    if matrix.get("data_disclaimer"):
        ws.cell(
            row=_SUBTITLE_ROW, column=1, value=matrix["data_disclaimer"]
        ).font = Font(size=10, color="6B7280")
    if matrix.get("scale_disclaimer"):
        ws.cell(
            row=_SUBTITLE_ROW + 1, column=1, value=matrix["scale_disclaimer"]
        ).font = Font(size=10, color="6B7280")

    header_row = 6
    _set_header_cell(ws, header_row, 1, "Group")
    _set_header_cell(ws, header_row, 2, "Competence")
    _set_header_cell(ws, header_row, 3, "Target")
    # Manager columns first, then AI columns — recruiters scan left-to-right.
    # The ✅ marker flags candidates whose AI run saw a transcript, so a
    # reader can tell a resume-only column from a fully-evidenced one.
    col = 4
    manager_first_col = col
    for cand in candidates:
        _set_header_cell(ws, header_row, col, f"Mgr {cand.get('name')}")
        col += 1
    ai_first_col = col
    for cand in candidates:
        suffix = " ✅" if cand.get("ai_full_data") else ""
        _set_header_cell(ws, header_row, col, f"AI {cand.get('name')}{suffix}")
        col += 1
    ai_last_col = col - 1
    ws.row_dimensions[header_row].height = 30

    body_start = header_row + 1
    row_idx = body_start
    # Track contiguous runs of the same group so the Group column can be
    # merged per run. Merging column A is safe: the Total row only sums
    # the candidate columns.
    group_runs: list[tuple[str, int, int]] = []
    for comp in competences:
        group = comp.get("group") or "—"
        if group_runs and group_runs[-1][0] == group:
            name, start, _ = group_runs[-1]
            group_runs[-1] = (name, start, row_idx)
        else:
            group_runs.append((group, row_idx, row_idx))

        _set_body_cell(ws, row_idx, 1, group, fill=_GROUP_FILL, bold=True)
        _set_body_cell(ws, row_idx, 2, comp.get("name"))
        _set_body_cell(
            ws,
            row_idx,
            3,
            _criticality_label(comp.get("criticality")),
            align=_CENTER_ALIGN,
        )

        cells: list[dict] = comp.get("cells") or []
        cell_by_cand = {str(c.get("candidate_id")): c for c in cells}
        # Manager values
        for offset, cand in enumerate(candidates):
            entry = cell_by_cand.get(str(cand.get("id"))) or {}
            ms = entry.get("manager_score")
            diverges = bool(entry.get("divergence"))
            tone = _AMBER_FILL if diverges else None
            if ms is None:
                text: Any = "—"
                fmt: str | None = None
            else:
                text = round(float(ms), 1)
                fmt = _DIVERGENCE_FORMAT if diverges else None
            _set_body_cell(
                ws,
                row_idx,
                manager_first_col + offset,
                text,
                align=_CENTER_ALIGN,
                fill=tone,
                number_format=fmt,
            )
        # AI values
        for offset, cand in enumerate(candidates):
            entry = cell_by_cand.get(str(cand.get("id"))) or {}
            ai_score = entry.get("ai_score")
            ai_status = (entry.get("ai_status") or "").lower()
            diverges = bool(entry.get("divergence"))
            tone = _AMBER_FILL if diverges else None
            ai_fmt: str | None = None
            if ai_status == "not_covered":
                ai_text: Any = "n/a"
            elif ai_score is None:
                ai_text = "—"
            else:
                ai_text = round(float(ai_score), 1)
                ai_fmt = _DIVERGENCE_FORMAT if diverges else None
            _set_body_cell(
                ws,
                row_idx,
                ai_first_col + offset,
                ai_text,
                align=_CENTER_ALIGN,
                fill=tone,
                number_format=ai_fmt,
            )
        row_idx += 1

    for _group, start, end in group_runs:
        if end > start:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
            top = ws.cell(row=start, column=1)
            top.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            # Bottom border under the last row of the run — the visual
            # separator between groups the spec asks for.
            for c in range(1, ai_last_col + 1):
                cell = ws.cell(row=end, column=c)
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=Side(style="medium", color="94A3B8"),
                )

    # Total row — a live SUM per candidate column, formatted as the spec's
    # "39/44 (89%)" so the denominator (which differs per AI column once
    # not-covered competences drop out) is visible in the cell itself.
    total_row = row_idx
    _set_body_cell(ws, total_row, 1, "", fill=_TOTAL_FILL)
    _set_body_cell(
        ws,
        total_row,
        2,
        f"TOTAL (max {round(max_score * len(competences), 1)})",
        fill=_TOTAL_FILL,
        bold=True,
    )
    _set_body_cell(ws, total_row, 3, "", fill=_TOTAL_FILL)
    if competences:
        body_first = body_start
        body_last = total_row - 1
        for offset, cand in enumerate(candidates):
            for col_idx, denominator, percent in (
                (
                    manager_first_col + offset,
                    cand.get("manager_denominator"),
                    cand.get("manager_percent"),
                ),
                (
                    ai_first_col + offset,
                    cand.get("ai_denominator"),
                    cand.get("ai_percent"),
                ),
            ):
                col_letter = get_column_letter(col_idx)
                span = f"{col_letter}{body_first}:{col_letter}{body_last}"
                denom = float(denominator or 0)
                if denom > 0:
                    formula = (
                        f'=SUM({span})&"/{_trim_number(denom)} ("'
                        f"&ROUND(SUM({span})/{_trim_number(denom)}*100,0)"
                        f'&"%)"'
                    )
                else:
                    formula = f"=SUM({span})"
                cell = ws.cell(row=total_row, column=col_idx, value=formula)
                cell.alignment = _CENTER_ALIGN
                cell.fill = _total_fill_for(percent)
                cell.font = _BOLD
                cell.border = _THIN_BORDER

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 10
    for col_idx in range(manager_first_col, ai_last_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


def _criticality_label(value: Any) -> str:
    """Abbreviated target level for the Matrix "Target" column."""
    lowered = str(value or "").strip().lower()
    if lowered in {"critical", "crit", "must"}:
        return "Crit"
    if lowered in {"important", "imp", "should"}:
        return "Imp"
    if lowered in {"desirable", "desired", "nice_to_have", "des", "could"}:
        return "Des"
    return "—"


def _trim_number(value: float) -> str:
    """Render 44.0 as ``44`` but keep 43.5 intact — the denominator ends
    up inside an Excel formula string, so trailing ``.0`` would leak into
    the displayed total."""
    return str(int(value)) if float(value).is_integer() else str(value)


def _total_fill_for(percent: Any) -> PatternFill:
    """Threshold colouring for the Total row: >75% emerald, 50-75% amber,
    below 50% rose. An unscored column stays neutral."""
    if percent is None:
        return _TOTAL_FILL
    try:
        pct = float(percent)
    except (TypeError, ValueError):
        return _TOTAL_FILL
    if pct > 75:
        return _EMERALD_FILL
    if pct >= 50:
        return _AMBER_FILL
    return _ROSE_FILL


# ---------------------------------------------------------------------------
# Sheet 3 — Detail per candidate
# ---------------------------------------------------------------------------


def _render_detail(
    ws: Worksheet,
    detail: dict,
    *,
    audience: str,
    branding: dict | None,
) -> None:
    _embed_logo(ws, branding)
    name = detail.get("candidate_name") or "Candidate"
    ws.cell(row=_TITLE_ROW, column=1, value=f"Detailed analysis — {name}").font = Font(
        size=14, bold=True
    )
    subtitle_parts = [
        detail.get("position"),
        detail.get("status"),
        f"Report date: {_today()}",
    ]
    subtitle = " | ".join([s for s in subtitle_parts if s])
    ws.cell(row=_SUBTITLE_ROW, column=1, value=subtitle).font = Font(
        size=10, color="6B7280"
    )

    row = 5
    # Resume is a paragraph, not a table — merged across A:E so the text
    # is readable instead of being clipped by column A's width.
    row = _detail_paragraph(ws, row, "RESUME", detail.get("resume_summary") or "—")

    hide_findings_for_hm = audience.lower() == "hiring_manager"
    scores = detail.get("scores") or []
    if scores:
        # HM audience: drop AI free-text reasoning + citations from the
        # Scores table — both regularly carry the same kind of
        # candidate-side process narrative as ``process_findings`` does,
        # and shipping them unfiltered would defeat the HM audience swap.
        if hide_findings_for_hm:
            headers = ["Group", "Competence", "Manager score", "AI score"]
            score_rows = [
                [
                    s.get("competence_group") or "—",
                    s.get("competence_name"),
                    _format_score(s.get("manager_score")),
                    _format_score(s.get("ai_score")),
                ]
                for s in scores
            ]
        else:
            headers = [
                "Group",
                "Competence",
                "Manager score",
                "AI score",
                "Citations from transcript",
                "AI reasoning",
            ]
            score_rows = [
                [
                    s.get("competence_group") or "—",
                    s.get("competence_name"),
                    _format_score(s.get("manager_score")),
                    _format_score(s.get("ai_score")),
                    "\n".join(s.get("citations") or []),
                    s.get("reasoning") or "",
                ]
                for s in scores
            ]
        row = _detail_table(ws, row, "COMPETENCE SCORES", headers, score_rows)

    blind_spots = detail.get("blind_spots") or []
    if blind_spots:
        row = _detail_table(
            ws,
            row,
            "BLIND SPOTS",
            ["Competence", "Suggested question"],
            [[b.get("competence"), b.get("suggested_question")] for b in blind_spots],
        )

    findings = detail.get("process_findings") or []
    if findings:
        if hide_findings_for_hm:
            # HM audience: replace raw process-findings text with the
            # positive reframe so the candidate-side issues never reach
            # the manager unfiltered.
            rows = [
                [
                    f.get("finding_type") or "",
                    f.get("positive_reframe")
                    or "Recommendation for the next interview.",
                ]
                for f in findings
            ]
            row = _detail_table(
                ws,
                row,
                "PROCESS RECOMMENDATIONS FOR NEXT INTERVIEW",
                ["Topic", "Suggestion"],
                rows,
            )
        else:
            row = _detail_table(
                ws,
                row,
                "PROCESS FINDINGS",
                ["Type", "Severity", "Description"],
                [
                    [
                        f.get("finding_type") or "",
                        f.get("severity") or "",
                        f.get("description") or "",
                    ]
                    for f in findings
                ],
            )

    # Red flags are candidate-side concerns by definition — HM audience
    # never sees them; the verdict-level risk fields are also redacted
    # below for the same reason.
    red_flags = detail.get("red_flags") or []
    if red_flags and not hide_findings_for_hm:
        row = _detail_table(
            ws,
            row,
            "RED FLAGS",
            ["Type", "Severity", "Description"],
            [
                [
                    f.get("flag_type") or "",
                    f.get("severity") or "",
                    f.get("description") or "",
                ]
                for f in red_flags
            ],
        )

    verdict = detail.get("verdict") or {}
    if verdict:
        verdict_lines = [
            (
                "Verdict",
                verdict.get("verdict_summary") or verdict.get("verdict") or "—",
            ),
            ("Key strength", verdict.get("key_strength") or "—"),
        ]
        if not hide_findings_for_hm:
            verdict_lines.extend(
                [
                    ("Key risk", verdict.get("key_risk") or "—"),
                    ("Mitigation", verdict.get("risk_mitigation") or "—"),
                ]
            )
        verdict_lines.append(
            (
                "Recommendation",
                verdict.get("recommendation_for_next_step") or "—",
            )
        )
        row = _detail_kv(ws, row, "FINAL VERDICT", verdict_lines)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 46
    ws.column_dimensions["F"].width = 46


def _format_score(value: Any) -> str:
    if value is None:
        return "—"
    if isinstance(value, (int, float)):
        return f"{round(float(value), 1)}"
    return str(value)


# Detail sections span columns A:E so the banner reads as a divider and
# free text is not clipped by column A's width.
_DETAIL_SPAN = 5


def _detail_banner(ws: Worksheet, row: int, title: str) -> int:
    """Section divider — ``─── TITLE ───`` merged across A:E."""
    cell = ws.cell(row=row, column=1, value=f"─── {title} ───")
    cell.font = _BOLD
    cell.fill = _GROUP_FILL
    cell.alignment = _BODY_ALIGN
    cell.border = _THIN_BORDER
    for col_idx in range(2, _DETAIL_SPAN + 1):
        filler = ws.cell(row=row, column=col_idx)
        filler.fill = _GROUP_FILL
        filler.border = _THIN_BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_DETAIL_SPAN)
    return row + 1


def _detail_paragraph(ws: Worksheet, row: int, title: str, text: str) -> int:
    """Banner + one block of free text merged across A:E (HRP-525)."""
    row = _detail_banner(ws, row, title)
    cell = ws.cell(row=row, column=1, value=_coerce(text))
    cell.alignment = _BODY_ALIGN
    cell.border = _THIN_BORDER
    for col_idx in range(2, _DETAIL_SPAN + 1):
        ws.cell(row=row, column=col_idx).border = _THIN_BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_DETAIL_SPAN)
    return row + 2


def _detail_table(
    ws: Worksheet,
    row: int,
    title: str,
    headers: list[str],
    rows_data: list[list[Any]],
) -> int:
    row = _detail_banner(ws, row, title)
    for col_idx, value in enumerate(headers, start=1):
        _set_header_cell(ws, row, col_idx, value)
    row += 1
    for body_row in rows_data:
        for col_idx, value in enumerate(body_row, start=1):
            _set_body_cell(ws, row, col_idx, value)
        row += 1
    return row + 1


def _detail_kv(
    ws: Worksheet, row: int, title: str, items: list[tuple[str, Any]]
) -> int:
    row = _detail_banner(ws, row, title)
    for key, value in items:
        _set_body_cell(ws, row, 1, key, bold=True)
        _set_body_cell(ws, row, 2, value)
        # Verdict prose is long — give it the same A:E width the resume
        # block gets instead of clipping it at column B.
        for col_idx in range(3, _DETAIL_SPAN + 1):
            ws.cell(row=row, column=col_idx).border = _THIN_BORDER
        ws.merge_cells(
            start_row=row, start_column=2, end_row=row, end_column=_DETAIL_SPAN
        )
        row += 1
    return row + 1


# ---------------------------------------------------------------------------
# Sheet 4 — Incomplete data (always emitted)
# ---------------------------------------------------------------------------


def _render_incomplete(
    ws: Worksheet, items: list[dict], *, branding: dict | None
) -> None:
    _embed_logo(ws, branding)
    ws.cell(
        row=_TITLE_ROW, column=1, value="Candidates with incomplete data"
    ).font = Font(size=14, bold=True)

    headers = ["Candidate", "What's missing", "What to do"]
    header_row = 4
    for col, value in enumerate(headers, start=1):
        _set_header_cell(ws, header_row, col, value)
    ws.row_dimensions[header_row].height = 22

    if not items:
        cell = ws.cell(
            row=header_row + 1,
            column=1,
            value="All candidates have complete data — no action required.",
        )
        cell.alignment = _BODY_ALIGN
        cell.font = Font(italic=True, color="6B7280")
        ws.merge_cells(
            start_row=header_row + 1,
            start_column=1,
            end_row=header_row + 1,
            end_column=3,
        )
    else:
        for idx, item in enumerate(items, start=header_row + 1):
            _set_body_cell(ws, idx, 1, item.get("candidate_name"))
            _set_body_cell(ws, idx, 2, item.get("missing"))
            _set_body_cell(ws, idx, 3, item.get("action"))

    _autosize(ws, headers, default=28)


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def _autosize(ws: Worksheet, headers: list[str], default: int = 24) -> None:
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(len(header) + 4, default), 60
        )


# ---------------------------------------------------------------------------
# Assessment canvas export (HRP-510)
# ---------------------------------------------------------------------------


def render_canvas_xlsx(matrix: dict, *, vacancy_title: str) -> bytes:
    """Flatten the assessment-matrix payload into a one-sheet workbook.

    Mirrors what the fullscreen canvas shows: candidates down the rows,
    competences across the columns, Manager and AI side by side, plus a
    totals block. Reuses the report palette so both exports look like the
    same product.
    """

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Canvas"

    competences: list[dict] = matrix.get("competences") or []
    candidates: list[dict] = matrix.get("candidates") or []

    ws.cell(row=1, column=1, value=f"Assessment canvas — {vacancy_title}").font = Font(
        size=14, bold=True
    )
    ws.cell(
        row=2,
        column=1,
        value=(
            f"{len(candidates)} candidates × {len(competences)} competences | "
            f"Scale: {matrix.get('scale_name') or 'default'} "
            f"(max {matrix.get('max_score')}) | "
            f"Round: {matrix.get('round') or 'latest'} | "
            f"Divergence threshold: {matrix.get('divergence_threshold')}"
        ),
    ).font = Font(size=10, color="6B7280")

    header_row = 4
    _set_header_cell(ws, header_row, 1, "Candidate")
    _set_header_cell(ws, header_row, 2, "Source")
    for offset, comp in enumerate(competences):
        _set_header_cell(ws, header_row, 3 + offset, comp.get("name"))
    total_col = 3 + len(competences)
    _set_header_cell(ws, header_row, total_col, "Total")
    ws.row_dimensions[header_row].height = 30

    row_idx = header_row + 1
    for cand in candidates:
        cell_by_comp = {
            str(c.get("competence_id")): c for c in (cand.get("cells") or [])
        }
        # Two rows per candidate — Manager then AI — so a reader can scan
        # either side without the cell text running together.
        for source in ("manager", "ai"):
            _set_body_cell(
                ws,
                row_idx,
                1,
                cand.get("name") if source == "manager" else "",
                bold=source == "manager",
            )
            _set_body_cell(ws, row_idx, 2, "Manager" if source == "manager" else "AI")
            for offset, comp in enumerate(competences):
                entry = cell_by_comp.get(str(comp.get("id"))) or {}
                diverges = bool(entry.get("divergence"))
                if source == "manager":
                    score = entry.get("manager_score")
                    text: Any = "—" if score is None else round(float(score), 1)
                else:
                    status = (entry.get("ai_status") or "").lower()
                    score = entry.get("ai_score")
                    if status == "not_covered":
                        text = "n/a"
                    elif score is None:
                        text = "—"
                    else:
                        text = round(float(score), 1)
                if diverges and text not in {"—", "n/a"}:
                    text = f"{text} ⚠"
                _set_body_cell(
                    ws,
                    row_idx,
                    3 + offset,
                    text,
                    align=_CENTER_ALIGN,
                    fill=_AMBER_FILL if diverges else None,
                )
            percent = (
                cand.get("manager_percent")
                if source == "manager"
                else cand.get("ai_percent")
            )
            _set_body_cell(
                ws,
                row_idx,
                total_col,
                "—" if percent is None else f"{percent}%",
                align=_CENTER_ALIGN,
                fill=_total_fill_for(percent),
                bold=True,
            )
            row_idx += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    for offset in range(len(competences)):
        ws.column_dimensions[get_column_letter(3 + offset)].width = 16
    ws.column_dimensions[get_column_letter(total_col)].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
