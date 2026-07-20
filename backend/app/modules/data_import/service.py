import base64
import uuid
from datetime import date, datetime
from io import BytesIO

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_password
from app.modules.auth.models import User
from app.modules.data_import.models import ImportJob
from app.modules.dictionary.models import DictionaryItem
from app.modules.employee.models import Course, Education, Employee, WorkExperience

VALID_IMPORT_TYPES = {"employees", "dictionaries"}


async def start_import(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    user_id: uuid.UUID,
    import_type: str,
    file: UploadFile,
) -> dict:
    if import_type not in VALID_IMPORT_TYPES:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Invalid import type. Valid: {', '.join(VALID_IMPORT_TYPES)}",
        )

    data = await file.read()
    wb = load_workbook(BytesIO(data), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

    job = ImportJob(
        tenant_id=tenant_id,
        import_type=import_type,
        file_name=file.filename or "unnamed",
        total_rows=len(rows),
        initiated_by=user_id,
    )
    db.add(job)
    await db.commit()
    await db.refresh(job)

    # Queue Celery task for background processing
    from app.core.task_enqueue import enqueue_task
    from app.modules.data_import.tasks import run_import_task

    b64_data = base64.b64encode(data).decode()
    enqueue_task(
        run_import_task,
        str(job.id),
        str(tenant_id),
        str(user_id),
        import_type,
        b64_data,
        tenant_id=tenant_id,
        user_id=user_id,
        module="data_import",
        action=f"import_{import_type}",
    )

    return _job_to_dict(job)


def _parse_date(value) -> date | None:
    """Parse a date from Excel cell value (date object or string)."""
    if value is None:
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


async def _import_employees(
    db: AsyncSession, tenant_id: uuid.UUID, rows: list
) -> tuple[int, int, dict]:
    """Import employees from Excel rows.

    Required columns (0-4): email, first_name, last_name, position, hire_date
    Optional GF1 columns (5-8): work_exp_title, work_exp_role, work_exp_start, work_exp_end
    Optional GF2 columns (9-13): edu_institution, edu_degree, edu_field, edu_start, edu_end
    Optional GF2 columns (14-16): course_title, course_provider, course_completed_date
    """
    processed = 0
    error_count = 0
    errors: dict[str, list] = {"rows": []}

    for i, row in enumerate(rows, start=2):
        try:
            if len(row) < 5 or not row[0]:
                errors["rows"].append({"row": i, "error": "Missing required fields"})
                error_count += 1
                continue

            email, first_name, last_name, position, hire_date = (
                row[0],
                row[1],
                row[2],
                row[3],
                row[4],
            )

            # Create user if not exists
            existing = await db.execute(
                select(User).where(
                    User.email == str(email), User.tenant_id == tenant_id
                )
            )
            user = existing.scalar_one_or_none()
            if not user:
                user = User(
                    email=str(email),
                    password_hash=hash_password("changeme123"),
                    first_name=str(first_name),
                    last_name=str(last_name),
                    tenant_id=tenant_id,
                )
                db.add(user)
                await db.flush()

            # Create employee if not exists
            existing_emp = await db.execute(
                select(Employee).where(
                    Employee.user_id == user.id, Employee.tenant_id == tenant_id
                )
            )
            emp = existing_emp.scalar_one_or_none()
            if not emp:
                # Get or create Position from text
                position_title = str(position) if position else "Employee"
                position_id = None
                if position:
                    from app.modules.position.models import Position

                    pos_result = await db.execute(
                        select(Position).where(
                            Position.tenant_id == tenant_id,
                            Position.title == position_title,
                        )
                    )
                    pos = pos_result.scalar_one_or_none()
                    if not pos:
                        pos = Position(
                            tenant_id=tenant_id,
                            title=position_title,
                            source="manual",
                        )
                        db.add(pos)
                        await db.flush()
                    position_id = pos.id

                emp = Employee(
                    user_id=user.id,
                    tenant_id=tenant_id,
                    position_id=position_id,
                    position_title=position_title,
                    hire_date=hire_date,
                )
                db.add(emp)
                await db.flush()

            # GF1: Work experience (columns 5-8)
            work_title = row[5] if len(row) > 5 and row[5] else None
            if work_title:
                work_role = str(row[6]) if len(row) > 6 and row[6] else None
                work_start = _parse_date(row[7] if len(row) > 7 else None)
                work_end = _parse_date(row[8] if len(row) > 8 else None)
                if work_start:
                    db.add(
                        WorkExperience(
                            employee_id=emp.id,
                            tenant_id=tenant_id,
                            title=str(work_title),
                            role=work_role,
                            start_date=work_start,
                            end_date=work_end,
                        )
                    )

            # GF2: Education (columns 9-13)
            edu_institution = row[9] if len(row) > 9 and row[9] else None
            if edu_institution:
                edu_degree = str(row[10]) if len(row) > 10 and row[10] else "Bachelor"
                edu_field = str(row[11]) if len(row) > 11 and row[11] else ""
                edu_start = _parse_date(row[12] if len(row) > 12 else None)
                edu_end = _parse_date(row[13] if len(row) > 13 else None)
                if edu_start:
                    db.add(
                        Education(
                            employee_id=emp.id,
                            tenant_id=tenant_id,
                            institution=str(edu_institution),
                            degree=edu_degree,
                            field_of_study=edu_field,
                            start_date=edu_start,
                            end_date=edu_end,
                        )
                    )

            # GF2: Courses (columns 14-16)
            course_title = row[14] if len(row) > 14 and row[14] else None
            if course_title:
                course_provider = str(row[15]) if len(row) > 15 and row[15] else None
                course_date = _parse_date(row[16] if len(row) > 16 else None)
                db.add(
                    Course(
                        employee_id=emp.id,
                        tenant_id=tenant_id,
                        title=str(course_title),
                        provider=course_provider,
                        completed_date=course_date,
                    )
                )

            processed += 1
        except Exception as e:  # noqa: BLE001 - per-row import isolation, recorded
            errors["rows"].append({"row": i, "error": str(e)})
            error_count += 1

    return processed, error_count, errors


async def _import_dictionaries(
    db: AsyncSession, tenant_id: uuid.UUID, rows: list
) -> tuple[int, int, dict]:
    """Import dictionaries from Excel rows: [type, title, description]"""
    processed = 0
    error_count = 0
    errors: dict[str, list] = {"rows": []}

    for i, row in enumerate(rows, start=2):
        try:
            if len(row) < 2 or not row[0] or not row[1]:
                errors["rows"].append({"row": i, "error": "Missing type or title"})
                error_count += 1
                continue

            item_type, title = str(row[0]), str(row[1])
            description = str(row[2]) if len(row) > 2 and row[2] else None

            # Check if exists
            existing = await db.execute(
                select(DictionaryItem).where(
                    DictionaryItem.type == item_type,
                    DictionaryItem.title == title,
                    DictionaryItem.tenant_id == tenant_id,
                )
            )
            if not existing.scalar_one_or_none():
                db.add(
                    DictionaryItem(
                        type=item_type,
                        title=title,
                        description=description,
                        tenant_id=tenant_id,
                    )
                )
            processed += 1
        except Exception as e:  # noqa: BLE001 - per-row import isolation, recorded
            errors["rows"].append({"row": i, "error": str(e)})
            error_count += 1

    return processed, error_count, errors


async def get_job(db: AsyncSession, tenant_id: uuid.UUID, job_id: uuid.UUID) -> dict:
    job = await db.get(ImportJob, job_id)
    if not job or job.tenant_id != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Import job not found")
    return _job_to_dict(job)


async def list_jobs(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    skip: int = 0,
    limit: int = 20,
) -> tuple[list[dict], int]:
    count_q = select(func.count(ImportJob.id)).where(ImportJob.tenant_id == tenant_id)
    total = (await db.execute(count_q)).scalar() or 0

    result = await db.execute(
        select(ImportJob)
        .where(ImportJob.tenant_id == tenant_id)
        .order_by(ImportJob.created_at.desc())
        .offset(skip)
        .limit(limit)
    )
    return [_job_to_dict(j) for j in result.scalars().all()], total


def _job_to_dict(j: ImportJob) -> dict:
    return {
        "id": j.id,
        "import_type": j.import_type,
        "file_name": j.file_name,
        "status": j.status,
        "total_rows": j.total_rows,
        "processed_rows": j.processed_rows,
        "error_rows": j.error_rows,
        "errors": j.errors,
        "initiated_by": j.initiated_by,
        "tenant_id": j.tenant_id,
        "started_at": j.started_at,
        "finished_at": j.finished_at,
        "created_at": j.created_at,
    }
