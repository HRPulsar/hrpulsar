"""Celery tasks for the data-import module.

Moved out of ``app.core.tasks`` (project-review #30) so core no longer
imports domain models: the bulk employee/dictionary import reads and
writes ``auth``, ``employee`` and ``dictionary`` rows and is enqueued
from ``data_import/service.py``.
"""

import logging

from app.core.celery_app import celery

logger = logging.getLogger(__name__)


@celery.task(bind=True, max_retries=1, default_retry_delay=30)
def run_import_task(
    self,
    job_id: str,
    tenant_id: str,
    user_id: str,
    import_type: str,
    b64_file_data: str,
) -> None:
    """Process data import in background. Updates ImportJob status as it runs."""
    import base64
    import uuid
    from datetime import date as date_type
    from datetime import datetime, timezone
    from io import BytesIO

    from openpyxl import load_workbook
    from sqlalchemy import select
    from sqlalchemy.orm import Session

    from app.config import settings
    from app.core.security import hash_password
    from app.database import make_sync_engine
    from app.modules.auth.models import User
    from app.modules.data_import.models import ImportJob
    from app.modules.dictionary.models import DictionaryItem
    from app.modules.employee.models import Course, Education, Employee, WorkExperience

    engine = make_sync_engine(settings.database_url)

    def _parse_date(value) -> date_type | None:
        if value is None:
            return None
        if isinstance(value, date_type):
            return value if not isinstance(value, datetime) else value.date()
        try:
            return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return None

    try:
        with Session(engine) as db:
            job = db.get(ImportJob, uuid.UUID(job_id))
            if not job:
                logger.error("Import job %s not found", job_id)
                return

            job.status = "processing"
            job.started_at = datetime.now(timezone.utc)
            db.commit()

            # Parse Excel
            raw = base64.b64decode(b64_file_data)
            wb = load_workbook(BytesIO(raw), read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            tid = uuid.UUID(tenant_id)
            processed = 0
            error_count = 0
            errors: dict[str, list] = {"rows": []}

            if import_type == "employees":
                for i, row in enumerate(rows, start=2):
                    try:
                        if len(row) < 5 or not row[0]:
                            errors["rows"].append(
                                {"row": i, "error": "Missing required fields"}
                            )
                            error_count += 1
                            continue

                        email = str(row[0])
                        first_name, last_name = str(row[1]), str(row[2])
                        position = str(row[3])
                        hire_date = row[4]

                        existing = db.execute(
                            select(User).where(
                                User.email == email, User.tenant_id == tid
                            )
                        ).scalar_one_or_none()

                        if not existing:
                            existing = User(
                                email=email,
                                password_hash=hash_password("changeme123"),
                                first_name=first_name,
                                last_name=last_name,
                                tenant_id=tid,
                            )
                            db.add(existing)
                            db.flush()

                        emp = db.execute(
                            select(Employee).where(
                                Employee.user_id == existing.id,
                                Employee.tenant_id == tid,
                            )
                        ).scalar_one_or_none()

                        if not emp:
                            emp = Employee(
                                user_id=existing.id,
                                tenant_id=tid,
                                position=position,
                                hire_date=hire_date,
                            )
                            db.add(emp)
                            db.flush()

                        # GF1: Work experience (columns 5-8)
                        work_title = row[5] if len(row) > 5 and row[5] else None
                        if work_title:
                            work_role = str(row[6]) if len(row) > 6 and row[6] else None
                            work_start = _parse_date(row[7] if len(row) > 7 else None)
                            work_end = _parse_date(row[8] if len(row) > 8 else None)
                            if work_start:
                                db.add(
                                    WorkExperience(
                                        employee_id=emp.id,
                                        tenant_id=tid,
                                        title=str(work_title),
                                        role=work_role,
                                        start_date=work_start,
                                        end_date=work_end,
                                    )
                                )

                        # GF2: Education (columns 9-13)
                        edu_institution = row[9] if len(row) > 9 and row[9] else None
                        if edu_institution:
                            edu_degree = (
                                str(row[10])
                                if len(row) > 10 and row[10]
                                else "Bachelor"
                            )
                            edu_field = (
                                str(row[11]) if len(row) > 11 and row[11] else ""
                            )
                            edu_start = _parse_date(row[12] if len(row) > 12 else None)
                            edu_end = _parse_date(row[13] if len(row) > 13 else None)
                            if edu_start:
                                db.add(
                                    Education(
                                        employee_id=emp.id,
                                        tenant_id=tid,
                                        institution=str(edu_institution),
                                        degree=edu_degree,
                                        field_of_study=edu_field,
                                        start_date=edu_start,
                                        end_date=edu_end,
                                    )
                                )

                        # GF2: Courses (columns 14-16)
                        course_title = row[14] if len(row) > 14 and row[14] else None
                        if course_title:
                            course_provider = (
                                str(row[15]) if len(row) > 15 and row[15] else None
                            )
                            course_date = _parse_date(
                                row[16] if len(row) > 16 else None
                            )
                            db.add(
                                Course(
                                    employee_id=emp.id,
                                    tenant_id=tid,
                                    title=str(course_title),
                                    provider=course_provider,
                                    completed_date=course_date,
                                )
                            )

                        processed += 1
                    except Exception as e:  # noqa: BLE001 - per-row isolation
                        errors["rows"].append({"row": i, "error": str(e)})
                        error_count += 1

            elif import_type == "dictionaries":
                for i, row in enumerate(rows, start=2):
                    try:
                        if len(row) < 2 or not row[0] or not row[1]:
                            errors["rows"].append(
                                {"row": i, "error": "Missing type or title"}
                            )
                            error_count += 1
                            continue

                        item_type, title = str(row[0]), str(row[1])
                        description = str(row[2]) if len(row) > 2 and row[2] else None

                        existing_item = db.execute(
                            select(DictionaryItem).where(
                                DictionaryItem.type == item_type,
                                DictionaryItem.title == title,
                                DictionaryItem.tenant_id == tid,
                            )
                        ).scalar_one_or_none()

                        if not existing_item:
                            db.add(
                                DictionaryItem(
                                    type=item_type,
                                    title=title,
                                    description=description,
                                    tenant_id=tid,
                                )
                            )
                        processed += 1
                    except Exception as e:  # noqa: BLE001 - per-row isolation
                        errors["rows"].append({"row": i, "error": str(e)})
                        error_count += 1

            job.processed_rows = processed
            job.error_rows = error_count
            job.errors = errors if error_count > 0 else None
            job.status = "completed"
            job.finished_at = datetime.now(timezone.utc)
            db.commit()
            logger.info(
                "Import job %s completed: %d processed, %d errors",
                job_id,
                processed,
                error_count,
            )

    except Exception as exc:
        logger.exception("Import job %s failed: %s", job_id, exc)
        try:
            with Session(engine) as db:
                job = db.get(ImportJob, uuid.UUID(job_id))
                if job:
                    job.status = "failed"
                    job.errors = {"rows": [], "fatal": str(exc)}
                    job.finished_at = datetime.now(timezone.utc)
                    db.commit()
        except Exception:
            logger.exception("Failed to update import job status to failed")
        raise self.retry(exc=exc)
    finally:
        engine.dispose()
